"""Εξαγωγή σε xlsx: ένα φύλλο με τις κατεδαφίσεις, ένα pivot ανά έτος/δήμο."""

from collections import Counter
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("ΑΔΑ", "ada", 16),
    ("Ημ/νία έκδοσης", "date", 14),
    ("Έτος", "year", 7),
    ("Είδος", "eidos", 22),
    ("Έκταση", "ektasi", 14),
    ("Δήμος", "dimos", 26),
    ("Δημοτική Ενότητα/Περιοχή", "dim_enotita", 24),
    ("Πόλη/Οικισμός", "poli", 20),
    ("Οδός", "odos", 24),
    ("Αρ. από", "ar_apo", 9),
    ("Αρ. έως", "ar_eos", 9),
    ("ΟΤ", "ot", 8),
    ("ΚΑΕΚ", "kaek", 16),
    ("Περιγραφή έργου", "perigrafi", 60),
    ("Όροφοι", "orofoi", 8),
    ("Lat", "lat", 11),
    ("Lon", "lon", 11),
    ("Ακρίβεια συντ/νων", "precision", 14),
    ("Ανάλυση PDF", "parse_ok", 12),
    ("PDF", "pdf_path", 8),
    ("Έλεγχος", "flags", 24),
]


def write_xlsx(rows, out_path):
    wb = Workbook()
    # σταθερά μεταδεδομένα — να μη διαρρέει όνομα χρήστη/περιβάλλοντος
    wb.properties.creator = "demolitions"
    wb.properties.lastModifiedBy = "demolitions"

    ws = wb.active
    ws.title = "Κατεδαφίσεις"
    bold = Font(bold=True)
    for col, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
        ws.column_dimensions[get_column_letter(col)].width = width
    for r, row in enumerate(rows, 2):
        for col, (_, key, _) in enumerate(COLUMNS, 1):
            value = row.get(key)
            if key == "parse_ok":
                value = "ΝΑΙ" if value else "ΟΧΙ"
            elif key == "date":
                value = date.fromisoformat(value)
            if key == "pdf_path":
                # σύνδεσμος στο PDF της Διαύγειας (προβολή, όχι λήψη)· ποτέ
                # τοπική διαδρομή — το xlsx μοιράζεται/ανεβαίνει αυτόνομα
                cell = ws.cell(row=r, column=col, value="PDF")
                cell.hyperlink = f"https://diavgeia.gov.gr/doc/{row['ada']}?inline=true"
                cell.font = Font(color="0563C1", underline="single")
                continue
            cell = ws.cell(row=r, column=col, value=value)
            if key == "date":
                cell.number_format = "DD/MM/YYYY"
            if key == "ada":
                cell.hyperlink = row["url"]
                cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 2)}"

    # pivot αυτοτελών αδειών (συγκρίσιμο μεταξύ ερευνών)· οι ενιαίες
    # οικοδομικές-με-κατεδάφιση, αν υπάρχουν, σε δικό τους φύλλο
    pure = [r for r in rows if r.get("eidos", "κατεδάφιση") == "κατεδάφιση"]
    bundled = [r for r in rows if r.get("eidos", "κατεδάφιση") != "κατεδάφιση"]
    _pivot_sheet(wb, "Ανά έτος-δήμο", pure, bold)
    if bundled:
        _pivot_sheet(wb, "Οικοδομικές με κατεδάφιση", bundled, bold)

    wb.save(out_path)


def _pivot_sheet(wb, title, rows, bold):
    pivot = Counter((row["year"], row["dimos"]) for row in rows)
    years = sorted({y for y, _ in pivot})
    dimoi = sorted({d for _, d in pivot})
    ws2 = wb.create_sheet(title)
    ws2.cell(row=1, column=1, value="Έτος").font = bold
    for c, d in enumerate(dimoi, 2):
        ws2.cell(row=1, column=c, value=d).font = bold
        ws2.column_dimensions[get_column_letter(c)].width = max(14, len(d) + 2)
    ws2.cell(row=1, column=len(dimoi) + 2, value="Σύνολο").font = bold
    for r, y in enumerate(years, 2):
        ws2.cell(row=r, column=1, value=y).font = bold
        for c, d in enumerate(dimoi, 2):
            ws2.cell(row=r, column=c, value=pivot.get((y, d), 0))
        ws2.cell(row=r, column=len(dimoi) + 2,
                 value=sum(pivot.get((y, d), 0) for d in dimoi))
    total_row = len(years) + 2
    ws2.cell(row=total_row, column=1, value="Σύνολο").font = bold
    for c, d in enumerate(dimoi, 2):
        ws2.cell(row=total_row, column=c,
                 value=sum(pivot.get((y, d), 0) for y in years))
    ws2.cell(row=total_row, column=len(dimoi) + 2, value=len(rows)).font = bold
