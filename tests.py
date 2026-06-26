#!/usr/bin/env python3
"""Δοκιμές χωρίς δίκτυο: python3 tests.py

Χρειάζονται μόνο το cache/kallikratis.json (υπάρχει μετά το πρώτο run).
"""

import io
import json
import os
import tempfile
import unittest
import zipfile
from datetime import datetime, timezone
from pathlib import Path

CACHE = str(Path(__file__).parent / "cache")

from demolitions.areas import (AreaError, list_areas, municipality_labels,
                               normalize, resolve_area)
from demolitions.diavgeia import (KIND_KATEDAFISI, KIND_OIKODOMIKI,
                                   _search_query, issue_date, permit_kind)
from demolitions.egsa87 import egsa87_to_wgs84
from demolitions.geocode import (PE_CENTROIDS, _poli_variants, _strip_dimos,
                                  row_out_of_region)
from demolitions.greek import dimos_display, greek_title, pretty_area
from demolitions.output import COLUMNS, write_xlsx
from demolitions.pdfparse import (_clean, _pdf_url, detect_extent, detect_floors,
                                   extract_polygon, is_nonbuilding, parse_fields)


class TestNormalize(unittest.TestCase):
    def test_accents_case_hyphens(self):
        self.assertEqual(normalize("Δήμος Δράμας"), "ΔΗΜΟΣ ΔΡΑΜΑΣ")
        self.assertEqual(normalize("ΚΕΑΣ - ΚΥΘΝΟΥ"), "ΚΕΑΣ ΚΥΘΝΟΥ")
        self.assertEqual(normalize("  πολλά   κενά "), "ΠΟΛΛΑ ΚΕΝΑ")
        self.assertEqual(normalize("Αχαΐας"), "ΑΧΑΙΑΣ")


class TestResolveArea(unittest.TestCase):
    def test_dimos(self):
        label, munis = resolve_area("Δήμος Δράμας", CACHE)
        self.assertEqual(label, "ΔΗΜΟΣ ΔΡΑΜΑΣ")
        self.assertEqual(list(munis), ["0201"])

    def test_nomos_vs_pe(self):
        _, nomos = resolve_area("Νομός Καβάλας", CACHE)
        _, pe = resolve_area("ΠΕ Καβάλας", CACHE)
        # ο νομός περιλαμβάνει και τη Θάσο (πρόθεμα 04), η ΠΕ όχι
        self.assertTrue(set(pe) < set(nomos))
        self.assertTrue(any(c.startswith("04") for c in nomos))
        self.assertFalse(any(c.startswith("04") for c in pe))

    def test_perifereia_kai_ellada(self):
        _, kriti = resolve_area("Περιφέρεια Κρήτης", CACHE)
        self.assertEqual(len(kriti), 24)
        _, ellada = resolve_area("Ελλάδα", CACHE)
        self.assertEqual(len(ellada), 326)

    def test_polles_perioxes(self):
        label, munis = resolve_area("Δήμος Δράμας, Δήμος Θάσου", CACHE)
        self.assertEqual(len(munis), 2)
        self.assertIn(",", label)

    def test_amfisimi_kai_agnosti(self):
        # μοναδικό διπλώνυμο ζευγάρι δήμων στην Ελλάδα
        with self.assertRaises(AreaError):
            resolve_area("Δήμος Ηρακλείου", CACHE)
        with self.assertRaises(AreaError):
            resolve_area("Δήμος Ασγκαμπάτ", CACHE)
        with self.assertRaises(AreaError):
            resolve_area("  ", CACHE)

    def test_omonimoi_dimoi_me_prosdiorismo(self):
        _, kriti = resolve_area("Δήμος Ηρακλείου (Κρήτης)", CACHE)
        _, attiki = resolve_area("Δήμος Ηρακλείου (Αττικής)", CACHE)
        self.assertEqual(list(kriti)[0][:2], "71")
        self.assertEqual(list(attiki)[0][:2], "46")

    def test_ola_ta_autocomplete_labels_epiluontai(self):
        """Κάθε επιλογή του autocomplete πρέπει να γίνεται δεκτή."""
        for a in list_areas(CACHE):
            label, munis = resolve_area(a["label"], CACHE)
            self.assertTrue(munis, f"κενή επίλυση: {a['label']}")

    def test_omonimoi_dimoi_exoun_diaforetikes_etiketes(self):
        labels = municipality_labels(("4604", "7101"), CACHE)
        self.assertEqual(labels["4604"]["display"], "Δήμος Ηρακλείου (Αττικής)")
        self.assertEqual(labels["7101"]["display"], "Δήμος Ηρακλείου (Κρήτης)")


class TestGreek(unittest.TestCase):
    def test_lexiko(self):
        self.assertEqual(greek_title("ΝΕΑ ΑΜΙΣΟΣ"), "Νέα Αμισός")
        self.assertEqual(greek_title("28ΗΣ ΟΚΤΩΒΡΙΟΥ"), "28ης Οκτωβρίου")
        self.assertEqual(greek_title("ΕΘΝΙΚΗΣ ΑΜΥΝΗΣ"), "Εθνικής Αμύνης")

    def test_agnostes_lexeis_xoris_tono(self):
        self.assertEqual(greek_title("ΧΩΡΙΣΤΗ"), "Χωριστή")  # στο λεξικό
        self.assertEqual(greek_title("ΑΓΝΩΣΤΟΧΩΡΙ"), "Αγνωστοχωρι")

    def test_teliko_sigma_kai_latinika(self):
        self.assertEqual(greek_title("ΛΙΜΕΝΑΣ ΘΑΣΟΥ"), "Λιμένας Θάσου")
        self.assertEqual(greek_title("OK 12"), "Ok 12")
        self.assertEqual(greek_title(""), "")

    def test_idempotent(self):
        for s in ("Νέα Αμισός", "Δήμος Δράμας", "Υψηλάντου 12"):
            self.assertEqual(greek_title(s), s)

    def test_dimoi(self):
        self.assertEqual(dimos_display("ΔΗΜΟΣ ΔΡΑΜΑΣ"), "Δήμος Δράμας")
        self.assertEqual(dimos_display("ΔΗΜΟΣ ΑΓΙΟΥ ΔΗΜΗΤΡΙΟΥ"),
                         "Δήμος Αγίου Δημητρίου")

    def test_pretty_area(self):
        self.assertEqual(pretty_area("ΔΗΜΟΣ ΔΡΑΜΑΣ, ΠΕΡΙΦΕΡΕΙΑ ΚΡΗΤΗΣ"),
                         "Δήμος Δράμας, Περιφέρεια Κρήτης")


class TestDetectFloors(unittest.TestCase):
    CASES = [
        ("ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ ΚΑΤΟΙΚΙΑΣ", 2),
        ("ΚΑΤΕΔΑΦΙΣΗ ΙΣΟΓΕΙΑΣ ΑΠΟΘΗΚΗΣ", 1),
        ("ΚΑΤΕΔΑΦΙΣΗ ΤΡΙΟΡΩΦΟΥ ΚΤΙΡΙΟΥ", 3),       # ορθογραφικό
        ("ΚΑΤΕΔΑΦΙΣΗ ΚΑΤΟΙΚΙΑΣ 2 ΟΡΟΦΩΝ", 2),
        ("ΚΑΤΕΔΑΦΙΣΗ ΚΤΙΡΙΟΥ ΤΡΙΩΝ ΟΡΟΦΩΝ", 3),
        ("ΙΣΟΓΕΙΑ ΚΑΤΟΙΚΙΑ ΚΑΙ ΔΙΩΡΟΦΗ ΑΠΟΘΗΚΗ", 2),  # max
        ("ΚΑΤΕΔΑΦΙΣΗ ΣΤΟ ΟΤ 15", None),               # όχι αριθμός ΟΤ
        ("ΚΑΤΕΔΑΦΙΣΗ ΚΑΤΟΙΚΙΑΣ", None),
    ]

    def test_cases(self):
        for desc, want in self.CASES:
            self.assertEqual(detect_floors(normalize(desc)), want, desc)


class TestNonBuilding(unittest.TestCase):
    def test_walls_fences_etc(self):
        for desc in ("ΚΑΤΕΔΑΦΙΣΗ ΤΟΙΧΙΟΥ", "ΚΑΤΕΔΑΦΙΣΗ ΜΑΝΤΡΟΤΟΙΧΟΥ",
                     "ΚΑΤΕΔΑΦΙΣΗ ΠΕΡΙΦΡΑΞΗΣ ΟΙΚΟΠΕΔΟΥ",
                     "ΚΑΤΕΔΑΦΙΣΗ ΚΟΛΥΜΒΗΤΙΚΗΣ ΔΕΞΑΜΕΝΗΣ"):
            self.assertTrue(is_nonbuilding(normalize(desc)), desc)

    def test_buildings_not_flagged(self):
        for desc in ("ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ ΚΑΤΟΙΚΙΑΣ",
                     "ΚΑΤΕΔΑΦΙΣΗ ΙΣΟΓΕΙΑΣ ΑΠΟΘΗΚΗΣ",
                     "ΚΑΤΕΔΑΦΙΣΗ ΚΑΤΟΙΚΙΑΣ ΚΑΙ ΠΕΡΙΦΡΑΞΗΣ",   # έχει κτίριο
                     "ΚΑΤΕΔΑΦΙΣΗ ΚΤΙΡΙΟΥ"):
            self.assertFalse(is_nonbuilding(normalize(desc)), desc)


class TestDetectExtent(unittest.TestCase):
    def test_cases(self):
        for desc, want in [
            ("ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ ΚΑΤΟΙΚΙΑΣ", "ολική"),
            ("ΚΑΤΕΔΑΦΙΣΗ ΤΜΗΜΑΤΟΣ ΚΤΙΡΙΟΥ", "τμηματική/μερική"),
            ("ΜΕΡΙΚΗ ΚΑΤΕΔΑΦΙΣΗ ΙΣΟΓΕΙΟΥ", "τμηματική/μερική"),
            ("ΤΜΗΜΑΤΙΚΗ ΚΑΤΕΔΑΦΙΣΗ ΑΠΟΘΗΚΗΣ", "τμηματική/μερική"),
        ]:
            self.assertEqual(detect_extent(normalize(desc)), want, desc)


class TestParseFields(unittest.TestCase):
    # απόσπασμα όπως το βγάζει το pdftotext -layout από τη φόρμα e-Άδειες
    SAMPLE = """\
Στοιχεία διεύθυνσης
  Οδός                 ΥΨΗΛΑΝΤΟΥ
  Αρ. από              12
  Πόλη/Οικισμός        ΔΡΑΜΑ
  Δήμος                ΔΡΑΜΑΣ
  Δημοτική Ενότητα /   ΝΕΑ ΑΜΙΣΟΣ
  ΟΤ                   45
  ΚΑΕΚ                 -
  Περιγραφή έργου      ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ ΚΑΤΟΙΚΙΑΣ
\fΔεύτερη σελίδα
  Οδός                 ΑΛΛΗ ΟΔΟΣ ΠΟΥ ΔΕΝ ΜΕΤΡΑΕΙ
"""

    def test_fields(self):
        f = parse_fields(self.SAMPLE)
        self.assertEqual(f["odos"], "ΥΨΗΛΑΝΤΟΥ")
        self.assertEqual(f["ar_apo"], "12")
        self.assertEqual(f["poli"], "ΔΡΑΜΑ")
        self.assertEqual(f["dim_enotita"], "ΝΕΑ ΑΜΙΣΟΣ")
        self.assertEqual(f["kaek"], "")          # παύλα -> κενό
        self.assertEqual(f["perigrafi"], "ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ ΚΑΤΟΙΚΙΑΣ")

    def test_proti_selida_mono(self):
        f = parse_fields(self.SAMPLE)
        self.assertNotIn("ΑΛΛΗ", f["odos"])

    def test_clean(self):
        self.assertEqual(_clean(" - "), "")
        self.assertEqual(_clean("τιμή "), "τιμή")


class TestPermitKind(unittest.TestCase):
    def test_katedafisi(self):
        self.assertEqual(permit_kind(
            "Άδεια Κατεδάφισης (ν.4759/2020): ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ"),
            KIND_KATEDAFISI)

    def test_oikodomiki_me_katedafisi(self):
        self.assertEqual(permit_kind(
            "Οικοδομική Άδεια (ν.4759/2020): ΑΔΕΙΑ ΚΑΤΕΔΑΦΙΣΗΣ & ΑΝΕΓΕΡΣΗ "
            "ΝΕΟΥ ΔΙΩΡΟΦΟΥ"), KIND_OIKODOMIKI)
        self.assertEqual(permit_kind(
            "Οικοδομική άδεια Κατηγορίας 1 χωρίς προέγκριση: Κατεδάφιση και "
            "ανέγερση"), KIND_OIKODOMIKI)
        self.assertEqual(permit_kind(
            "Οικοδομική Άδεια (ν.4759/2020): ΚΑΤΕΔΑΦΙΣΗ BARBECUE, ΑΝΕΓΕΡΣΗ ΝΕΑΣ "
            "ΙΣΟΓΕΙΟΥ ΟΙΚΟΔΟΜΗΣ"), KIND_OIKODOMIKI)

    def test_aporriptontai(self):
        for s in ("Προέγκριση Άδειας Κατεδάφισης: ...",
                  "Αναθεώρηση Άδειας Κατεδάφισης: ...",
                  "Ενημέρωση Οικοδομικής Άδειας: ΚΑΤΕΔΑΦΙΣΗ ...",
                  "Προέγκριση Οικοδομικής Άδειας: ΚΑΤΕΔΑΦΙΣΗ ...",
                  "Οικοδομική Άδεια (ν.4759/2020): ΑΝΕΓΕΡΣΗ ΚΑΤΟΙΚΙΑΣ",
                  "Οικοδομική Άδεια (ν.4759/2020): ΚΑΤΑΣΚΕΥΗ ΜΟΝΑΔΑΣ "
                  "ΔΙΑΧΕΙΡΙΣΗΣ ΑΠΟΒΛΗΤΩΝ ΕΚΣΚΑΦΩΝ, ΚΑΤΑΣΚΕΥΩΝ ΚΑΙ ΚΑΤΕΔΑΦΙΣΕΩΝ",
                  "Έγκριση Εκτέλεσης Εργασιών: ΚΑΤΕΔΑΦΙΣΗ ΕΠΙΚΙΝΔΥΝΟΥ"):
            self.assertIsNone(permit_kind(s), s)

    def test_dotted_code_in_description_does_not_break_classification(self):
        # κωδικός με τελείες ΜΕΣΑ στην περιγραφή (μετά το «:») — δεν επηρεάζει
        # την επικεφαλίδα, άρα η ταξινόμηση παραμένει σωστή (πραγματικά subjects)
        self.assertEqual(permit_kind(
            "Άδεια Κατεδάφισης (ν.4759/2020): ΑΔΕΙΑ ΚΑΤΕΔΑΦΙΣΗΣ ΓΙΑ ΤΑ "
            "ΚΤΙΡΙΑ 5.0 - 5.0.1 - 5.1 - 7.3."), KIND_KATEDAFISI)
        self.assertEqual(permit_kind(
            "Οικοδομική Άδεια (ν.4759/2020): ΚΑΤΕΔΑΦΙΣΗ ΚΤΙΡΙΟΥ 6.2 ΚΑΙ "
            "ΑΝΕΓΕΡΣΗ"), KIND_OIKODOMIKI)

    def test_excluded_kinds_with_dotted_code_stay_excluded(self):
        # οι αποκλεισμένες πράξεις παραμένουν αποκλεισμένες, ακόμη κι όταν η
        # περιγραφή περιέχει κωδικό με τελείες (πραγματικά subjects)
        for s in (
            "Έγκριση Εκτέλεσης Εργασιών: ΚΑΤΕΔΑΦΙΣΗ ΕΠΙΚΙΝΔΥΝΟΥ ΚΤΙΣΜΑΤΟΣ "
            "ΒΑΣΕΙ ΤΗΣ ΥΠ' ΑΡΙΘΜΟΝ 4087-3.7.2024",
            "Προέγκριση Οικοδομικής Αδείας (ν.4759/2020): ΚΑΤΕΔΑΦΙΣΗ ΠΑΛΑΙΩΝ "
            "ΚΤΙΣΜΑΤΩΝ & ΑΝΕΓΕΡΣΗ",
            "Ενημέρωση Άδειας Κατεδάφισης (ν.4759/2020): ΚΑΤΕΔΑΦΙΣΗ",
            "Αναθεώρηση Άδειας Κατεδάφισης χωρίς μεταβολή κάλυψης/δόμησης: "
            "ΚΑΤΕΔΑΦΙΣΗ",
        ):
            self.assertIsNone(permit_kind(s), s)

    def test_leading_code_prefix_does_not_occur_in_real_data(self):
        # Τεκμηρίωση επιβεβαιωμένης παραδοχής: subject με κωδικό-πρόθεμα πριν τη
        # λέξη-είδος ΔΕΝ υπάρχει στα πραγματικά δεδομένα της Διαύγειας (έλεγχος
        # σε 16.823 subjects). Καρφώνουμε την ΤΡΕΧΟΥΣΑ συμπεριφορά: επιστρέφει
        # None, αφού η επικεφαλίδα δεν ξεκινά με «ΑΔΕΙΑ ΚΑΤΕΔΑΦΙΣΗΣ».
        self.assertIsNone(permit_kind(
            "6.4.6.1 Άδεια Κατεδάφισης: ΚΑΤΕΔΑΦΙΣΗ ΔΙΩΡΟΦΗΣ"))

    def test_search_query_is_broad_enough_for_bundled_permits(self):
        q = _search_query("2024-01-01", "2024-06-30")
        self.assertIn('subject:"Κατεδάφιση"', q)
        self.assertNotIn('subject:"Άδεια Κατεδάφισης"', q)


class TestIssueDate(unittest.TestCase):
    def test_issue_date_uses_greece_timezone(self):
        ts = int(datetime(2024, 4, 16, 21, 5, 11, tzinfo=timezone.utc).timestamp() * 1000)
        self.assertEqual(issue_date({"issueDate": ts}).isoformat(), "2024-04-17")

    def test_issue_date_midday_is_stable(self):
        ts = int(datetime(2024, 4, 16, 12, 0, 0, tzinfo=timezone.utc).timestamp() * 1000)
        self.assertEqual(issue_date({"issueDate": ts}).isoformat(), "2024-04-16")


class TestEgsa87(unittest.TestCase):
    def test_drama(self):
        lat, lon = egsa87_to_wgs84(512801.149, 4555388.78)
        self.assertAlmostEqual(lat, 41.152291, places=5)
        self.assertAlmostEqual(lon, 24.154344, places=5)

    def test_athina(self):
        lat, lon = egsa87_to_wgs84(476000, 4203000)   # κέντρο Αθήνας
        self.assertAlmostEqual(lat, 37.9769, places=3)
        self.assertAlmostEqual(lon, 23.7284, places=3)

    def test_ektos_orion(self):
        self.assertIsNone(egsa87_to_wgs84(0, 0))
        self.assertIsNone(egsa87_to_wgs84(512801, 9999999))


class TestExtractPolygon(unittest.TestCase):
    SAMPLE = """\
Στοιχεία
Συντεταγμένες   512801.1493522985 4555388.77952756,512817.2889679111
                4555393.674329015,512821.78689357365 4555375.153458641,512806.04415375483
                4555370.655532978




                                                              Σελίδα 3 από 4
"""

    def test_polygon(self):
        poly = extract_polygon(self.SAMPLE)
        self.assertEqual(len(poly), 4)
        for lat, lon in poly:
            self.assertAlmostEqual(lat, 41.152, places=2)
            self.assertAlmostEqual(lon, 24.154, places=2)

    def test_xoris_pedio(self):
        self.assertIsNone(extract_polygon("Άλλο κείμενο χωρίς συντεταγμένες"))

    def test_long_polygon(self):
        # 15 κορυφές, μία ανά γραμμή — δεν πρέπει να κόβονται (παλιό όριο: 7)
        pairs = [f"{512800 + i}.0 {4555300 + i}.0" for i in range(15)]
        text = ("Συντεταγμένες   " + pairs[0] + ",\n"
                + "".join(f"                {p},\n" for p in pairs[1:-1])
                + f"                {pairs[-1]}\n\n            Σελίδα 3 από 4\n")
        poly = extract_polygon(text)
        self.assertEqual(len(poly), 15)


class TestPdfText(unittest.TestCase):
    def test_pdf_url_ignores_untrusted_document_url(self):
        self.assertEqual(_pdf_url("ΑΔΑ123"), "https://diavgeia.gov.gr/doc/ΑΔΑ123")

    def test_missing_binary_returns_none(self):
        from unittest import mock
        from demolitions import pdfparse
        with mock.patch.object(pdfparse.subprocess, "run",
                               side_effect=FileNotFoundError):
            self.assertIsNone(pdfparse.pdf_text("/nonexistent.pdf"))

    def test_nonzero_exit_returns_none(self):
        from unittest import mock
        from demolitions import pdfparse
        fake = mock.Mock(returncode=1, stdout=b"")
        with mock.patch.object(pdfparse.subprocess, "run", return_value=fake):
            self.assertIsNone(pdfparse.pdf_text("/x.pdf"))


class _FakeStreamResponse:
    """Εικονική streaming απόκριση requests: υποστηρίζει context manager και
    iter_content, αλλά το .content σκάει — ώστε το test να αποδεικνύει ότι το
    download_pdf ΔΕΝ υλοποιεί ολόκληρο το σώμα στη μνήμη."""

    def __init__(self, body, chunk=65536):
        self._body, self._chunk = body, chunk

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False

    def raise_for_status(self):
        pass

    def iter_content(self, chunk_size):
        for i in range(0, len(self._body), chunk_size):
            yield self._body[i:i + chunk_size]

    @property
    def content(self):
        raise AssertionError("το download_pdf δεν πρέπει να αγγίζει το .content")


class TestDownloadPdf(unittest.TestCase):
    def test_streams_to_disk_without_loading_content(self):
        from unittest import mock
        from demolitions import pdfparse
        body = b"%PDF-1.4\n" + b"x" * (65536 * 2 + 17)   # >2 chunks
        decision = {"ada": "ΑΔΑ-STREAM"}
        with tempfile.TemporaryDirectory() as tmp:
            with mock.patch.object(
                    pdfparse.session, "get",
                    return_value=_FakeStreamResponse(body)) as get, \
                 mock.patch.object(pdfparse.time, "sleep"):
                path = pdfparse.download_pdf(decision, tmp)
            # (a) το PDF γράφτηκε σωστά, με τα ακριβή bytes
            self.assertIsNotNone(path)
            self.assertEqual(Path(path).read_bytes(), body)
            # (b) η λήψη ζητήθηκε με stream=True
            self.assertTrue(get.called)
            self.assertEqual(get.call_args.kwargs.get("stream"), True)
            # (c) το .content (πλήρες σώμα) δεν προσπελάστηκε — αλλιώς ο fake
            # θα είχε σκάσει με AssertionError πριν φτάσουμε εδώ

    def test_non_pdf_leaves_no_file_and_returns_none(self):
        from unittest import mock
        from demolitions import pdfparse
        decision = {"ada": "ΑΔΑ-HTML"}
        with tempfile.TemporaryDirectory() as tmp:
            with mock.patch.object(
                    pdfparse.session, "get",
                    return_value=_FakeStreamResponse(b"<html>not a pdf</html>")):
                path = pdfparse.download_pdf(decision, tmp)
            self.assertIsNone(path)
            self.assertEqual(
                list((Path(tmp) / "pdf").glob("*.pdf")), [])   # χωρίς υπόλειμμα


class TestGeocodeHelpers(unittest.TestCase):
    def test_strip_dimos(self):
        self.assertEqual(_strip_dimos("Δήμος Δράμας"), "Δράμας")
        self.assertEqual(_strip_dimos("ΔΗΜΟΣ ΔΡΑΜΑΣ"), "Δραμας")
        self.assertEqual(_strip_dimos("Δήμος Ηρακλείου (Αττικής)"), "Ηρακλείου")

    def test_poli_variants(self):
        self.assertEqual(_poli_variants("Οικισμός Ποταμιας Θάσου", "Θάσου"),
                         ["Ποταμιας", "Ποταμιας Θάσου", "Οικισμός Ποταμιας Θάσου"])
        self.assertEqual(_poli_variants("Δράμα/Προαστειο", "Δράμας"),
                         ["Δράμα/Προαστειο", "Δράμα", "Προαστειο"])
        self.assertEqual(_poli_variants("Δράμα", "Δράμας"), ["Δράμα"])
        self.assertEqual(_poli_variants("", "Δράμας"), [])

    def test_transient_failures_do_not_become_cached_misses(self):
        from unittest import mock
        from demolitions import geocode
        with tempfile.TemporaryDirectory() as tmp:
            g = geocode.Geocoder(tmp)
            with mock.patch.object(geocode.session, "get",
                                   side_effect=geocode.requests.RequestException), \
                 mock.patch.object(geocode.time, "sleep"):
                self.assertIsNone(g._query("Δράμα, Ελλάδα"))
                self.assertIsNone(g._query("Δράμα, Ελλάδα"))
            self.assertEqual(g.cache, {})

    def test_empty_hits_cache_are_reused(self):
        from unittest import mock
        from demolitions import geocode
        fake = mock.Mock(ok=True)
        fake.json.return_value = []
        with tempfile.TemporaryDirectory() as tmp:
            g = geocode.Geocoder(tmp)
            with mock.patch.object(geocode.session, "get", return_value=fake) as get, \
                 mock.patch.object(geocode.time, "sleep"):
                self.assertIsNone(g._query("Άγνωστο μέρος, Ελλάδα"))
                self.assertIsNone(g._query("Άγνωστο μέρος, Ελλάδα"))
            self.assertEqual(get.call_count, 1)
            self.assertEqual(g.cache["Άγνωστο μέρος, Ελλάδα"]["status"], "miss")


class TestOutput(unittest.TestCase):
    def test_xlsx(self):
        from openpyxl import load_workbook
        rows = [
            {"ada": "ΑΔΑ1", "url": "https://diavgeia.gov.gr/decision/view/ΑΔΑ1",
             "date": "2021-01-15", "year": 2021, "dimos": "Δήμος Δράμας",
             "dim_enotita": "", "poli": "Δράμα", "odos": "Υψηλάντου",
             "ar_apo": "12", "ar_eos": "", "ot": "", "kaek": "",
             "perigrafi": "ΚΑΤΕΔΑΦΙΣΗ", "orofoi": 2, "lat": 41.15, "lon": 24.14,
             "precision": "οδός", "parse_ok": True,
             "pdf_path": "pdf/Δήμος Δράμας/2021/ΑΔΑ1.pdf", "flags": ""},
            {"ada": "ΑΔΑ2", "url": "https://diavgeia.gov.gr/decision/view/ΑΔΑ2",
             "date": "2022-03-01", "year": 2022, "dimos": "Δήμος Θάσου",
             "dim_enotita": "", "poli": "", "odos": "", "ar_apo": "",
             "ar_eos": "", "ot": "", "kaek": "", "perigrafi": "",
             "orofoi": None, "lat": None, "lon": None, "precision": "",
             "parse_ok": False, "pdf_path": "", "flags": "πιθανό διπλό"},
        ]
        col = {key: i for i, (_, key, _) in enumerate(COLUMNS, 1)}
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            write_xlsx(rows, path)
            wb = load_workbook(path)
            ws = wb["Κατεδαφίσεις"]
            self.assertEqual(ws.max_row, 3)
            self.assertEqual(ws.cell(2, col["ada"]).hyperlink.target,
                             "https://diavgeia.gov.gr/decision/view/ΑΔΑ1")
            self.assertEqual(ws.cell(2, col["date"]).number_format, "DD/MM/YYYY")
            # η στήλη PDF δείχνει στη Διαύγεια (προβολή), ποτέ τοπική διαδρομή
            self.assertEqual(ws.cell(2, col["pdf_path"]).hyperlink.target,
                             "https://diavgeia.gov.gr/doc/ΑΔΑ1?inline=true")
            self.assertEqual(ws.cell(3, col["pdf_path"]).hyperlink.target,
                             "https://diavgeia.gov.gr/doc/ΑΔΑ2?inline=true")
            self.assertEqual(ws.cell(3, col["parse_ok"]).value, "ΟΧΙ")
            self.assertEqual(wb.properties.creator, "demolitions")
            pivot = wb["Ανά έτος-δήμο"]
            self.assertEqual(pivot.cell(1, 2).value, "Δήμος Δράμας")
            self.assertEqual(pivot.cell(4, 4).value, 2)      # γενικό σύνολο
            self.assertNotIn("Οικοδομικές με κατεδάφιση", wb.sheetnames)

    def test_xlsx_me_oikodomikes(self):
        from openpyxl import load_workbook
        base = {"ada": "Χ", "url": "u", "dim_enotita": "", "poli": "",
                "odos": "", "ar_apo": "", "ar_eos": "", "ot": "", "kaek": "",
                "perigrafi": "", "orofoi": None, "lat": None, "lon": None,
                "precision": "", "parse_ok": True, "pdf_path": "", "flags": ""}
        rows = [
            {**base, "date": "2021-01-01", "year": 2021,
             "dimos": "Δήμος Δράμας", "eidos": "κατεδάφιση"},
            {**base, "date": "2021-02-01", "year": 2021,
             "dimos": "Δήμος Δράμας", "eidos": "οικοδομική με κατεδάφιση"},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            write_xlsx(rows, path)
            wb = load_workbook(path)
            self.assertIn("Οικοδομικές με κατεδάφιση", wb.sheetnames)
            # το βασικό pivot μετρά μόνο τις αυτοτελείς
            self.assertEqual(wb["Ανά έτος-δήμο"].cell(3, 3).value, 1)
            self.assertEqual(wb["Οικοδομικές με κατεδάφιση"].cell(3, 3).value, 1)


class TestRunsConsistency(unittest.TestCase):
    """Έλεγχοι ακεραιότητας στα αποθηκευμένα runs (αν υπάρχουν)."""

    def test_rows_json(self):
        for run_dir in (Path(__file__).parent / "runs").glob("*"):
            f = run_dir / "rows.json"
            if not f.exists():
                continue
            rows = json.loads(f.read_text(encoding="utf-8"))
            m = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
            self.assertEqual(m["n_rows"], len(rows), run_dir.name)
            for r in rows:
                self.assertIn("decision/view", r["url"], run_dir.name)
                # τα PDF υπάρχουν στον δίσκο μόνο αν δεν έχουν εκκαθαριστεί
                if r["pdf_path"] and m.get("has_pdfs", True):
                    self.assertTrue((run_dir / r["pdf_path"]).exists(),
                                    f"{run_dir.name}: λείπει {r['pdf_path']}")


class TestRunPipelinePdfCallback(unittest.TestCase):
    """Ο pdf_callback ανεβάζει/σβήνει κάθε PDF αμέσως — έτσι ο εφήμερος
    δίσκος δεν γεμίζει σε μεγάλες αναζητήσεις (το bug της Αττικής 2025)."""

    def _decision(self, i):
        ts = int(datetime(2021, 1, 15, 12, tzinfo=timezone.utc).timestamp() * 1000)
        return {"ada": f"ΑΔΑ{i}", "issueDate": ts,
                "subject": "Άδεια Κατεδάφισης (ν.4759/2020): ΚΑΤΕΔΑΦΙΣΗ ΚΑΤΟΙΚΙΑΣ",
                "extraFieldValues": {"municipality": "0201"}}  # Δήμος Δράμας

    def _fake_parse(self, decision, cache_dir):
        return {"ada": decision["ada"],
                "url": f"https://diavgeia.gov.gr/decision/view/{decision['ada']}",
                "perigrafi": "ΚΑΤΕΔΑΦΙΣΗ ΚΑΤΟΙΚΙΑΣ", "parse_ok": True,
                "odos": "Οδός", "ar_apo": str(decision["ada"][-1]), "ar_eos": "",
                "poli": "Δράμα", "dim_enotita": "", "ot": "", "kaek": "",
                "orofoi": 1, "ektasi": "ολική", "nonbuilding": False}

    def _run(self, tmp, n_decisions, n_cached, calls, **kwargs):
        from unittest import mock
        from demolitions import pipeline
        cache_dir = Path(tmp) / "cache"
        (cache_dir / "pdf").mkdir(parents=True)
        decisions = [self._decision(i) for i in range(n_decisions)]
        for i in range(n_cached):     # μόνο τα πρώτα n_cached έχουν cached PDF
            (cache_dir / "pdf" / f"ΑΔΑ{i}.pdf").write_bytes(b"%PDF-1.4 x" * 50)
        from datetime import date as _date
        with mock.patch.object(pipeline, "search_permits", return_value=decisions), \
             mock.patch.object(pipeline, "parse_decision", side_effect=self._fake_parse):
            return pipeline.run_pipeline(
                "Δήμος Δράμας", _date(2021, 1, 1), _date(2021, 12, 31),
                Path(tmp) / "run", cache_dir=cache_dir,
                log=lambda m: None,
                pdf_callback=lambda dest, rel: calls.append((dest, rel)),
                **kwargs)

    def test_callback_fires_once_per_cached_pdf(self):
        with tempfile.TemporaryDirectory() as tmp:
            calls = []
            result = self._run(tmp, n_decisions=3, n_cached=2, calls=calls)
            # 3 άδειες, 2 με cached PDF -> 2 κλήσεις
            self.assertEqual(len(calls), 2)
            for dest, rel in calls:
                self.assertTrue(rel.startswith("pdf/"))
                # το PDF υπάρχει στον δίσκο τη στιγμή του callback (το upload
                # το διαβάζει· ο R2 backend το σβήνει αμέσως μετά)
                self.assertTrue(Path(dest).is_file())
                self.assertEqual(str(Path(dest).relative_to(result.run_dir)), rel)
            # η άδεια χωρίς cached PDF έχει κενό pdf_path, χωρίς callback
            blank = [r for r in result.rows if not r["pdf_path"]]
            self.assertEqual(len(blank), 1)
            self.assertEqual(blank[0]["ada"], "ΑΔΑ2")

    def test_no_callback_when_not_provided(self):
        # χωρίς pdf_callback (CLI / τοπικό) το pipeline τρέχει κανονικά
        with tempfile.TemporaryDirectory() as tmp:
            from unittest import mock
            from demolitions import pipeline
            from datetime import date as _date
            cache_dir = Path(tmp) / "cache"
            (cache_dir / "pdf").mkdir(parents=True)
            (cache_dir / "pdf" / "ΑΔΑ0.pdf").write_bytes(b"%PDF-1.4 x")
            with mock.patch.object(pipeline, "search_permits",
                                   return_value=[self._decision(0)]), \
                 mock.patch.object(pipeline, "parse_decision",
                                   side_effect=self._fake_parse):
                result = pipeline.run_pipeline(
                    "Δήμος Δράμας", _date(2021, 1, 1), _date(2021, 12, 31),
                    Path(tmp) / "run", cache_dir=cache_dir, log=lambda m: None)
            self.assertEqual(result.rows[0]["pdf_path"],
                             "pdf/Δήμος Δράμας/2021/ΑΔΑ0.pdf")

    def test_pdf_staged_during_download_phase_not_after(self):
        """Η αντιγραφή/ανέβασμα γίνεται ΜΕΣΑ στη φάση download (overlap), όχι
        σε χωριστό δεύτερο loop — αλλιώς χάνεται η επικάλυψη upload/download."""
        from unittest import mock
        from demolitions import pipeline
        from datetime import date as _date
        with tempfile.TemporaryDirectory() as tmp:
            cache_dir = Path(tmp) / "cache"
            (cache_dir / "pdf").mkdir(parents=True)
            decisions = [self._decision(i) for i in range(3)]
            for i in range(3):
                (cache_dir / "pdf" / f"ΑΔΑ{i}.pdf").write_bytes(b"%PDF-1.4 x")
            events = []

            def parse(d, c):
                events.append(("parse", d["ada"]))
                return self._fake_parse(d, c)

            with mock.patch.object(pipeline, "search_permits", return_value=decisions), \
                 mock.patch.object(pipeline, "parse_decision", side_effect=parse):
                pipeline.run_pipeline(
                    "Δήμος Δράμας", _date(2021, 1, 1), _date(2021, 12, 31),
                    Path(tmp) / "run", cache_dir=cache_dir, log=lambda m: None,
                    pdf_callback=lambda dest, rel: events.append(
                        ("upload", Path(dest).stem)))

            first_upload = next(k for k, e in enumerate(events) if e[0] == "upload")
            last_parse = max(k for k, e in enumerate(events) if e[0] == "parse")
            # το πρώτο ανέβασμα πρέπει να συμβεί ΠΡΙΝ τελειώσουν όλα τα parse
            self.assertLess(first_upload, last_parse,
                            f"χωρίς overlap upload/download: {events}")

    def test_free_cache_deletes_cache_copy_after_staging(self):
        """Με free_cache=True (hosted) το PDF σβήνεται από την cache μόλις
        ανέβει — αλλιώς ο εφήμερος δίσκος γεμίζει με όλα τα PDF του run."""
        from unittest import mock
        from demolitions import pipeline
        from datetime import date as _date
        with tempfile.TemporaryDirectory() as tmp:
            cache_dir = Path(tmp) / "cache"
            (cache_dir / "pdf").mkdir(parents=True)
            (cache_dir / "pdf" / "ΑΔΑ0.pdf").write_bytes(b"%PDF-1.4 x")
            with mock.patch.object(pipeline, "search_permits",
                                   return_value=[self._decision(0)]), \
                 mock.patch.object(pipeline, "parse_decision",
                                   side_effect=self._fake_parse):
                pipeline.run_pipeline(
                    "Δήμος Δράμας", _date(2021, 1, 1), _date(2021, 12, 31),
                    Path(tmp) / "run", cache_dir=cache_dir, log=lambda m: None,
                    pdf_callback=lambda dest, rel: None, free_cache=True)
            # το PDF σβήστηκε από την cache, αλλά υπάρχει στο run_dir (staging)
            self.assertFalse((cache_dir / "pdf" / "ΑΔΑ0.pdf").exists())
            self.assertTrue((Path(tmp) / "run" / "pdf" / "Δήμος Δράμας" /
                             "2021" / "ΑΔΑ0.pdf").exists())

    def test_default_keeps_cache_copy(self):
        # default (τοπικό): η cache διατηρείται για ταχύτερα re-runs
        from unittest import mock
        from demolitions import pipeline
        from datetime import date as _date
        with tempfile.TemporaryDirectory() as tmp:
            cache_dir = Path(tmp) / "cache"
            (cache_dir / "pdf").mkdir(parents=True)
            (cache_dir / "pdf" / "ΑΔΑ0.pdf").write_bytes(b"%PDF-1.4 x")
            with mock.patch.object(pipeline, "search_permits",
                                   return_value=[self._decision(0)]), \
                 mock.patch.object(pipeline, "parse_decision",
                                   side_effect=self._fake_parse):
                pipeline.run_pipeline(
                    "Δήμος Δράμας", _date(2021, 1, 1), _date(2021, 12, 31),
                    Path(tmp) / "run", cache_dir=cache_dir, log=lambda m: None)
            self.assertTrue((cache_dir / "pdf" / "ΑΔΑ0.pdf").exists())

    def test_size_estimate_logged(self):
        with tempfile.TemporaryDirectory() as tmp:
            logs = []
            from unittest import mock
            from demolitions import pipeline
            from datetime import date as _date
            cache_dir = Path(tmp) / "cache"
            (cache_dir / "pdf").mkdir(parents=True)
            with mock.patch.object(pipeline, "search_permits",
                                   return_value=[self._decision(i) for i in range(2)]), \
                 mock.patch.object(pipeline, "parse_decision",
                                   side_effect=self._fake_parse):
                pipeline.run_pipeline(
                    "Δήμος Δράμας", _date(2021, 1, 1), _date(2021, 12, 31),
                    Path(tmp) / "run", cache_dir=cache_dir, log=logs.append)
            self.assertTrue(any("Εκτιμώμενο μέγεθος PDF" in m for m in logs))


class TestEnrichGeocodeOutOfRegion(unittest.TestCase):
    """enrich_geocode -> _flag_out_of_region: ολοκληρωμένος έλεγχος ότι η
    εγγραφή εκτός περιοχής σημαίνεται ΑΚΡΙΒΩΣ μία φορά (per-row, χωρίς bbox)."""

    def test_enrich_geocode_flags_out_of_area(self):
        """Κτίσμα Κρήτης χρεωμένο σε δήμο Αθηνών σημαίνεται· οι αττικές όχι.
        Χωρίς καμία κλήση Nominatim (όλες οι εγγραφές έχουν συντεταγμένες PDF)."""
        from demolitions import pipeline
        base = {
            "url": "https://diavgeia.gov.gr/decision/view/X",
            "dimos": "Δήμος Αθηναίων",
            "precision": "κτίσμα (PDF)", "flags": "", "date": "2021-01-15",
            "year": 2021, "muni_code": "4505", "eidos": "κατεδάφιση",
            "dim_enotita": "", "poli": "", "odos": "", "ar_apo": "",
            "ar_eos": "", "ot": "", "kaek": "", "perigrafi": "",
            "orofoi": None, "ektasi": "ολική", "nonbuilding": False,
            "parse_ok": True, "pdf_path": "", "dimos_pdf": "",
        }
        rows = [{**base, "ada": f"ΑΔΑ{i}", "lat": 38.0 + i * 0.02, "lon": 23.7}
                for i in range(10)]
        rows.append({**base, "ada": "ΚΡΗΤΗ", "lat": 35.3, "lon": 25.1})
        manifest = {"run_id": "test_run", "area_query": "Περιφέρεια Αττικής",
                    "area": "Αττική", "from": "2021-01-01", "to": "2021-12-31",
                    "created": "2021-12-31", "n_rows": len(rows), "geocoded": False,
                    "has_pdfs": False}
        with tempfile.TemporaryDirectory() as tmp:
            run_dir = Path(tmp) / "test_run"
            run_dir.mkdir()
            (run_dir / "rows.json").write_text(
                json.dumps(rows, ensure_ascii=False), "utf-8")
            (run_dir / "run.json").write_text(
                json.dumps(manifest, ensure_ascii=False), "utf-8")
            pipeline.enrich_geocode(run_dir, cache_dir=tmp, log=lambda m: None)
            result_rows = json.loads((run_dir / "rows.json").read_text("utf-8"))
        crete = next(r for r in result_rows if r["ada"] == "ΚΡΗΤΗ")
        self.assertIn("εκτός περιοχής αναζήτησης", crete["flags"])
        # ακριβώς ΕΝΑ flag «εκτός περιοχής» (όχι διπλό από δεύτερο πέρασμα)
        self.assertEqual(crete["flags"].count("εκτός περιοχής αναζήτησης"), 1,
                         crete["flags"])
        attica = [r for r in result_rows if r["ada"] != "ΚΡΗΤΗ"]
        for r in attica:
            self.assertNotIn("εκτός περιοχής", r["flags"],
                             f"false positive για {r['ada']}")


class TestRowOutOfRegion(unittest.TestCase):
    """row_out_of_region — έλεγχος ανά εγγραφή με βάση τον κωδικό ΠΕ.

    Άγκυρα είναι το 2ψήφιο πρόθεμα του muni_code (Περιφερειακή Ενότητα), όχι το
    αμφίσημο όνομα δήμου — άρα άτρωτο στις ομωνυμίες και στο ποσοστό μόλυνσης."""

    # κέντρα από το πακεταρισμένο gazetteer (offline, ντετερμινιστικά)
    ATTICA = (38.0, 23.7)        # ~ΠΕ 46 (Β. Τομέας Αθηνών)
    CRETE = (35.34, 25.15)       # πραγματικό κτίσμα Ηρακλείου Κρήτης
    RHODES = (36.43, 28.22)      # Νότιο Αιγαίο

    def _row(self, lat, lon, muni_code, precision="κτίσμα (PDF)"):
        return {"lat": lat, "lon": lon, "muni_code": muni_code,
                "precision": precision, "flags": ""}

    def test_bundled_gazetteer_covers_every_pe_prefix(self):
        # κάθε πρόθεμα του PREFIX_PE πρέπει να έχει κέντρο, εντός ορίων Ελλάδας
        from demolitions.areas import PREFIX_PE
        for prefix in PREFIX_PE:
            self.assertIn(prefix, PE_CENTROIDS, prefix)
        for prefix, (lat, lon) in PE_CENTROIDS.items():
            self.assertTrue(34.5 <= lat <= 42.0 and 19.0 <= lon <= 30.1,
                            f"{prefix}: ({lat},{lon}) εκτός Ελλάδας")

    def test_crete_building_tagged_to_attica_flagged(self):
        # το πραγματικό παράδειγμα: κτίσμα Κρήτης χρεωμένο στον δήμο Ηρακλείου
        # Αττικής (κωδικός 4604, ΠΕ 46) -> ~323 km -> σημαίνεται
        flag = row_out_of_region(self._row(*self.CRETE, muni_code="4604"))
        self.assertIsNotNone(flag)
        self.assertIn("εκτός περιοχής αναζήτησης", flag)
        self.assertIn("km", flag)

    def test_all_25_flagged_regardless_of_total_size(self):
        """Acceptance §8.1: και τα 25 Κρήτης σημαίνονται ανεξάρτητα από το
        ποσοστό μόλυνσης — συμπεριλαμβανομένων των περιπτώσεων 29–56% όπου το
        παλιό bbox+IQR σήμαινε 0/25."""
        import random
        rng = random.Random(0)

        def crete():
            return (35.34 + rng.uniform(-0.1, 0.1),
                    25.15 + rng.uniform(-0.1, 0.1))

        def attica():
            return (38.0 + rng.uniform(-0.4, 0.4),
                    23.7 + rng.uniform(-0.3, 0.3))

        for n_att in (1000, 100, 60, 40, 20):       # 2.4% .. 55.6% μόλυνση
            attica_rows = [self._row(*attica(), muni_code="4604")
                           for _ in range(n_att)]
            crete_rows = [self._row(*crete(), muni_code="4604")
                          for _ in range(25)]
            flagged = sum(1 for r in crete_rows if row_out_of_region(r))
            self.assertEqual(flagged, 25, f"n_att={n_att}: {flagged}/25")
            # καμία αττική εγγραφή δεν σημαίνεται (κανένα ψευδώς θετικό)
            fp = sum(1 for r in attica_rows if row_out_of_region(r))
            self.assertEqual(fp, 0, f"n_att={n_att}: {fp} ψευδώς θετικά")

    def test_generalized_non_crete_case(self):
        """Acceptance §8.2: γενικευμένο — δήμος Λάρισας (ηπειρωτική ΠΕ 22)
        αλλά κτίσμα στη Ρόδο (Νότιο Αιγαίο) -> σημαίνεται, χωρίς ειδική λογική
        Κρήτης/Ηρακλείου."""
        flag = row_out_of_region(self._row(*self.RHODES, muni_code="2201"))
        self.assertIsNotNone(flag)
        self.assertIn("εκτός περιοχής αναζήτησης", flag)

    def test_clean_single_region_no_false_positive(self):
        """Acceptance §8.3: καθαρή αναζήτηση μίας περιφέρειας — τίποτα δεν
        σημαίνεται."""
        import random
        rng = random.Random(3)
        for _ in range(50):
            r = self._row(38.0 + rng.uniform(-0.4, 0.4),
                          23.7 + rng.uniform(-0.3, 0.3), muni_code="4604")
            self.assertIsNone(row_out_of_region(r))

    def test_near_border_not_flagged(self):
        """Acceptance §8.3: εγγραφή στα όρια επιμήκους ΠΕ δεν υπερ-σημαίνεται.
        Ορεστιάδα (~143 km από το κέντρο της ΠΕ Έβρου) μένει εντός."""
        r = self._row(41.50, 26.53, muni_code="0303")   # Δήμος Ορεστιάδας
        self.assertIsNone(row_out_of_region(r))

    def test_low_precision_and_missing_data_ignored(self):
        # χωρίς αξιόπιστες συντεταγμένες ή χωρίς κωδικό -> δεν ελέγχεται
        self.assertIsNone(row_out_of_region(
            self._row(*self.CRETE, muni_code="4604", precision="δήμος")))
        self.assertIsNone(row_out_of_region(
            {"lat": None, "lon": None, "muni_code": "4604",
             "precision": "κτίσμα (PDF)", "flags": ""}))
        self.assertIsNone(row_out_of_region(
            self._row(*self.CRETE, muni_code="")))      # χωρίς κωδικό


class TestFlagOutOfRegionPipeline(unittest.TestCase):
    """_flag_out_of_region — code-anchored έλεγχος ανά εγγραφή στο pipeline."""

    def _base(self, muni_code):
        return {"url": "https://diavgeia.gov.gr/decision/view/X",
                "dimos": "Δ", "precision": "κτίσμα (PDF)",
                "flags": "", "date": "2021-01-15", "year": 2021,
                "muni_code": muni_code, "eidos": "κατεδάφιση",
                "dim_enotita": "", "poli": "", "odos": "", "ar_apo": "",
                "ar_eos": "", "ot": "", "kaek": "", "perigrafi": "",
                "orofoi": None, "ektasi": "ολική", "nonbuilding": False,
                "parse_ok": True, "pdf_path": "", "dimos_pdf": ""}

    def test_heavy_contamination_all_flagged(self):
        """§5/§8.1: 25 Κρήτης + 40 Αττικής (38% μόλυνση) — όλα σημαίνονται,
        καμία αττική εγγραφή δεν σημαίνεται ψευδώς."""
        import random
        from demolitions.pipeline import _flag_out_of_region
        rng = random.Random(7)
        rows = [{**self._base("4604"), "ada": f"ΑΤΤ{i}",
                 "lat": 38.0 + rng.uniform(-0.4, 0.4),
                 "lon": 23.7 + rng.uniform(-0.3, 0.3)} for i in range(40)]
        rows += [{**self._base("4604"), "ada": f"ΚΡΗ{i}",
                  "lat": 35.34 + rng.uniform(-0.1, 0.1),
                  "lon": 25.15 + rng.uniform(-0.1, 0.1)} for i in range(25)]
        n_out = _flag_out_of_region(rows)
        crete = [r for r in rows if r["ada"].startswith("ΚΡΗ")]
        attica = [r for r in rows if r["ada"].startswith("ΑΤΤ")]
        self.assertEqual(sum(1 for r in crete
                             if "εκτός περιοχής" in r["flags"]), 25)
        self.assertGreaterEqual(n_out, 25)
        for r in attica:
            self.assertNotIn("εκτός περιοχής", r["flags"], r["ada"])

    def test_whole_country_no_false_positives(self):
        """§8.3: αναζήτηση «Ελλάδα» — κάθε εγγραφή κοντά στη δική της ΠΕ· ένα
        απομακρυσμένο νησί ΔΕΝ σημαίνεται (ο per-row έλεγχος κρίνει κάθε εγγραφή
        ως προς τη δική της ΠΕ, χωρίς bbox πάνω σε όλο το result set)."""
        from demolitions.pipeline import _flag_out_of_region
        rows = [
            {**self._base("4604"), "ada": "ΑΘΗΝΑ", "lat": 38.0, "lon": 23.75},
            {**self._base("0701"), "ada": "ΘΕΣ", "lat": 40.64, "lon": 22.94},
            {**self._base("7101"), "ada": "ΗΡΑΚΛΕΙΟ", "lat": 35.34, "lon": 25.13},
            {**self._base("6901"), "ada": "ΡΟΔΟΣ", "lat": 36.43, "lon": 28.22},
            {**self._base("2001"), "ada": "ΗΓΟΥΜ", "lat": 39.50, "lon": 20.27},
            {**self._base("0201"), "ada": "ΔΡΑΜΑ", "lat": 41.15, "lon": 24.14},
        ]
        n_out = _flag_out_of_region(rows)
        self.assertEqual(n_out, 0)
        for r in rows:
            self.assertNotIn("εκτός περιοχής", r["flags"], r["ada"])

    def test_wide_single_region_aegean_no_false_positives(self):
        """Regression (review P1-1): «Περιφέρεια Νοτίου Αιγαίου» απλώνεται ~400
        km (Κυκλάδες -> Δωδεκάνησα). Οι νόμιμες απομακρυσμένες νήσοι (Ρόδος,
        Κως, Κάρπαθος, Καστελλόριζο), σωστά χρεωμένες στη δική τους ΠΕ, ΔΕΝ
        πρέπει να σημαίνονται ως «εκτός περιοχής» — το παλιό bbox τις σήμαινε."""
        import random
        from demolitions.pipeline import _flag_out_of_region
        rng = random.Random(11)
        # μάζα Κυκλάδων (θα έθετε εκεί το κέντρο ενός bbox)
        rows = [{**self._base("6701"), "ada": f"ΝΑΞ{i}",   # ΠΕ Νάξου
                 "lat": 37.10 + rng.uniform(-0.2, 0.2),
                 "lon": 25.38 + rng.uniform(-0.2, 0.2)} for i in range(40)]
        # νόμιμες απομακρυσμένες νήσοι, σωστός κωδικός ΠΕ
        far = [("6901", "ΡΟΔΟΣ", 36.43, 28.22),     # ΠΕ Ρόδου
               ("6401", "ΚΩΣ", 36.89, 27.29),        # ΠΕ Κω
               ("6201", "ΚΑΡΠΑΘΟΣ", 35.51, 27.21),   # ΠΕ Καρπάθου
               ("6901", "ΚΑΣΤΕΛ", 36.15, 29.59)]     # Καστελλόριζο (ΠΕ Ρόδου)
        for code, ada, lat, lon in far:
            rows.append({**self._base(code), "ada": ada, "lat": lat, "lon": lon})
        n_out = _flag_out_of_region(rows)
        self.assertEqual(n_out, 0, "ψευδώς θετικά σε ευρεία περιφέρεια Αιγαίου")
        for r in rows:
            self.assertNotIn("εκτός περιοχής", r["flags"], r["ada"])


class TestR2Storage(unittest.TestCase):
    """Έλεγχος του R2 backend με εικονικό S3 (moto), αν είναι διαθέσιμο."""

    def setUp(self):
        import importlib.util
        if importlib.util.find_spec("moto") is None:
            self.skipTest("moto δεν είναι εγκατεστημένο")

    def _make_staging(self, store, run_id, n_pdfs, pdf_bytes=b"%PDF-1.4 test",
                      created="2024-01-01", created_at=None):
        d = store.staging_dir(run_id)
        rows = []
        for i in range(n_pdfs):
            rel = f"pdf/Δήμος Χ/2021/ΑΔΑ{i}.pdf"
            p = d / rel
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(pdf_bytes)
            rows.append({"ada": f"ΑΔΑ{i}", "pdf_path": rel})
        (d / "rows.json").write_text(json.dumps(rows, ensure_ascii=False), "utf-8")
        (d / (run_id + ".xlsx")).write_bytes(b"xlsx-bytes")
        manifest = {"run_id": run_id, "n_rows": n_pdfs, "has_pdfs": True,
                    "created": created}
        if created_at:
            manifest["created_at"] = created_at
        (d / "run.json").write_text(json.dumps(
            manifest), "utf-8")

    def test_roundtrip_and_eviction(self):
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            import os
            os.environ.update(R2_BUCKET="demo-test-bucket", R2_ACCOUNT_ID="acc",
                              R2_ACCESS_KEY_ID="k", R2_SECRET_ACCESS_KEY="s")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s", pdf_cap=40,
                              endpoint_url="https://s3.amazonaws.com")

            # δύο run, 3 PDF το καθένα (13 bytes -> 39 bytes/run)
            self._make_staging(store, "old", 3, created="2024-01-01",
                               created_at="2024-01-01T10:00:00+00:00")
            store.save_run("old")
            self._make_staging(store, "new", 3, created="2024-01-02",
                               created_at="2024-01-02T10:00:00+00:00")
            store.save_run("new")

            self.assertTrue(store.exists("new"))
            self.assertEqual(len(store.list_runs()), 2)
            # ανάγνωση member
            gen, size = store.open_member("new", "new.xlsx")
            self.assertEqual(b"".join(gen), b"xlsx-bytes")
            self.assertEqual(len(list(store.iter_pdfs("new"))), 3)

            # free_local_pdfs: σβήνει τοπικά, το R2 μένει ανέπαφο
            store.free_local_pdfs("new")
            self.assertFalse((store.staging_dir("new") / "pdf").exists())
            self.assertEqual(len(list(store.iter_pdfs("new"))), 3)
            store.save_meta("new")   # ανέβασμα μόνο json/xlsx, χωρίς σφάλμα

            # cap=40: χωράει 1 run (39b), το παλιότερο χάνει τα PDF του
            store.enforce_pdf_cap()
            kept = {m["run_id"]: m.get("has_pdfs") for m in store.list_runs()}
            self.assertEqual(kept["new"], True)
            self.assertEqual(kept["old"], False)
            self.assertEqual(len(list(store.iter_pdfs("old"))), 0)
            # μεταδεδομένα του old παραμένουν
            self.assertEqual(store.read_manifest("old")["n_rows"], 3)

            store.delete_run("new")
            self.assertFalse(store.exists("new"))

    def test_upload_pdf_immediate_uploads_and_deletes_local(self):
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")
            d = store.staging_dir("r1")
            p = d / "pdf" / "Δήμος Χ" / "2021" / "ΑΔΑ1.pdf"
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(b"%PDF-1.4 test")
            relpath = str(p.relative_to(d))

            store.upload_pdf_immediate("r1", relpath, p)

            self.assertFalse(p.exists())                        # τοπικό αντίγραφο σβήστηκε
            self.assertEqual(len(list(store.iter_pdfs("r1"))), 1)  # αρχείο στο R2

    def test_cleanup_removes_partial_r2_uploads(self):
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")
            d = store.staging_dir("r1")
            p = d / "pdf" / "Δήμος Χ" / "2021" / "ΑΔΑ1.pdf"
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(b"%PDF-1.4 test")
            relpath = str(p.relative_to(d))

            store.upload_pdf_immediate("r1", relpath, p)
            self.assertEqual(len(list(store.iter_pdfs("r1"))), 1)
            self.assertFalse(store.exists("r1"))   # run.json δεν έχει ανέβει ακόμα

            store.cleanup("r1")

            # cleanup αφαιρεί partial R2 uploads όταν run.json λείπει
            self.assertEqual(len(list(store.iter_pdfs("r1"))), 0)

    def test_delete_orphans_sweeps_uploads_without_manifest(self):
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")
            # ολοκληρωμένο run (έχει run.json) — μένει
            self._make_staging(store, "done1", 2,
                               created_at="2024-01-01T10:00:00+00:00")
            store.save_run("done1")
            # orphan: PDF χωρίς run.json (run που σκοτώθηκε πριν το save_run)
            store.s3.put_object(Bucket="demo-test-bucket",
                                Key="runs/killed/pdf/a.pdf", Body=b"%PDF x")
            store.s3.put_object(Bucket="demo-test-bucket",
                                Key="runs/killed/pdf/b.pdf", Body=b"%PDF y")
            # active: ανεβαίνει ακόμη (χωρίς run.json) — προστατεύεται
            store.s3.put_object(Bucket="demo-test-bucket",
                                Key="runs/active/pdf/c.pdf", Body=b"%PDF z")

            freed = store.delete_orphans(keep_ids=("active",))
            self.assertEqual(freed, 1)
            self.assertTrue(store.exists("done1"))                  # έμεινε
            self.assertEqual(len(list(store.iter_pdfs("killed"))), 0)  # σβήστηκε
            self.assertEqual(len(list(store.iter_pdfs("active"))), 1)  # προστατεύθηκε

    def test_resilient_body_resumes_after_dropped_read(self):
        """Αν κοπεί η σύνδεση R2 στη μέση, το _resilient_body ξανανοίγει με
        Range και ολοκληρώνει — δεν κόβει το zip (αιτία του 2.3 αντί 3.24 GB)."""
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")
            data = b"abcdefgh" * 30000          # 240 KB -> πολλά chunks
            key = "runs/r1/pdf/a.pdf"
            store.s3.put_object(Bucket="demo-test-bucket", Key=key, Body=data)

            class FlakyBody:
                """Σκάει στο 2ο read· μετά συμπεριφέρεται κανονικά."""
                def __init__(self, real):
                    self.real, self.reads = real, 0
                def read(self, n=-1):
                    self.reads += 1
                    if self.reads == 2:
                        raise ConnectionError("dropped")
                    return self.real.read(n)
                def close(self):
                    self.real.close()

            first = store.s3.get_object(Bucket="demo-test-bucket", Key=key)["Body"]
            out = b"".join(store._resilient_body(key, FlakyBody(first), len(data)))
            self.assertEqual(out, data)         # πλήρες, παρά την πτώση

    def test_cache_roundtrip_survives_ephemeral_disk(self):
        """push_cache/pull_cache διατηρούν το geocode.json στο R2 ώστε να
        επιβιώνει spin-down (το /tmp σβήνεται)· το κλειδί είναι έξω από runs/."""
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")
            with tempfile.TemporaryDirectory() as c1, tempfile.TemporaryDirectory() as c2:
                # κανένα cache ακόμη -> pull επιστρέφει False
                self.assertFalse(store.pull_cache("geocode.json", c1))
                payload = '{"Δράμα, Ελλάδα": {"status": "hit", "result": [41.1, 24.1]}}'
                (Path(c1) / "geocode.json").write_text(payload, "utf-8")

                self.assertTrue(store.push_cache("geocode.json", c1))
                # νέος (εφήμερος) δίσκος c2: pull το ξαναφέρνει ατόφιο
                self.assertTrue(store.pull_cache("geocode.json", c2))
                self.assertEqual((Path(c2) / "geocode.json").read_text("utf-8"), payload)

            # το cache δεν θεωρείται run (έξω από runs/) -> ούτε orphan ούτε χρήση
            self.assertEqual(store.delete_orphans(), 0)
            self.assertEqual(store.list_runs(), [])
            self.assertEqual(store.usage()["storage_bytes"], 0)

    def test_push_cache_noop_when_file_absent(self):
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")
            with tempfile.TemporaryDirectory() as c:
                self.assertFalse(store.push_cache("geocode.json", c))

    def test_eviction_uses_creation_time_not_manifest_mtime(self):
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            os.environ.update(R2_BUCKET="demo-test-bucket", R2_ACCOUNT_ID="acc",
                              R2_ACCESS_KEY_ID="k", R2_SECRET_ACCESS_KEY="s")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s", pdf_cap=40,
                              endpoint_url="https://s3.amazonaws.com")

            self._make_staging(store, "old", 3, created="2024-01-01",
                               created_at="2024-01-01T10:00:00+00:00")
            store.save_run("old")
            self._make_staging(store, "new", 3, created="2024-01-02",
                               created_at="2024-01-02T10:00:00+00:00")
            store.save_run("new")

            manifest = store.read_manifest("old")
            manifest["geocoded"] = True
            store.write_manifest("old", manifest)  # παλιό bug: το έκανε «νεότερο»

            store.enforce_pdf_cap()
            kept = {m["run_id"]: m.get("has_pdfs") for m in store.list_runs()}
            self.assertEqual(kept["new"], True)
            self.assertEqual(kept["old"], False)

    def test_enforce_cap_reconciles_stale_has_pdfs(self):
        """has_pdfs=True αλλά τα PDF σβήστηκαν out-of-band (pdf_bytes==0) ->
        το enforce_pdf_cap διορθώνει τη σημαία σε False χωρίς να χάνει
        μεταδεδομένα (αυτο-θεραπεία στο επόμενο πέρασμα)."""
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s", pdf_cap=10 ** 9,
                              endpoint_url="https://s3.amazonaws.com")

            self._make_staging(store, "stale", 3,
                               created_at="2024-01-01T10:00:00+00:00")
            store.save_run("stale")
            # out-of-band διαγραφή των PDF· το run.json μένει με has_pdfs=True
            store.delete_pdfs("stale")
            self.assertTrue(store.read_manifest("stale")["has_pdfs"])  # μπαγιάτικο
            self.assertEqual(len(list(store.iter_pdfs("stale"))), 0)

            store.enforce_pdf_cap()

            m = store.read_manifest("stale")
            self.assertFalse(m["has_pdfs"])           # διορθώθηκε
            self.assertEqual(m["n_rows"], 3)          # μεταδεδομένα ανέπαφα
            self.assertEqual(len(list(store.iter_pdfs("stale"))), 0)

    def test_enforce_cap_keeps_valid_newest_reconciles_stale_oldest(self):
        """Δύο run, μεγάλο cap: το νεότερο με ανέπαφα PDF κρατά has_pdfs=True
        και τα αντικείμενά του· το παλαιότερο με out-of-band σβησμένα PDF
        διορθώνεται σε False — δεν θεωρείται «χωράει» ψευδώς."""
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s", pdf_cap=10 ** 9,
                              endpoint_url="https://s3.amazonaws.com")

            self._make_staging(store, "old", 3, created="2024-01-01",
                               created_at="2024-01-01T10:00:00+00:00")
            store.save_run("old")
            self._make_staging(store, "new", 3, created="2024-01-02",
                               created_at="2024-01-02T10:00:00+00:00")
            store.save_run("new")
            store.delete_pdfs("old")     # παλαιότερο: out-of-band διαγραφή

            store.enforce_pdf_cap()

            kept = {m["run_id"]: m.get("has_pdfs") for m in store.list_runs()}
            self.assertEqual(kept["new"], True)       # ανέπαφο νεότερο
            self.assertEqual(kept["old"], False)      # διορθώθηκε
            self.assertEqual(len(list(store.iter_pdfs("new"))), 3)
            self.assertEqual(len(list(store.iter_pdfs("old"))), 0)

    def test_enforce_cap_reconcile_is_idempotent(self):
        """Δεύτερο πέρασμα enforce_pdf_cap σε ήδη διορθωμένο stale run δεν
        ρίχνει σφάλμα και αφήνει has_pdfs=False."""
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s", pdf_cap=10 ** 9,
                              endpoint_url="https://s3.amazonaws.com")

            self._make_staging(store, "stale", 2,
                               created_at="2024-01-01T10:00:00+00:00")
            store.save_run("stale")
            store.delete_pdfs("stale")

            store.enforce_pdf_cap()
            self.assertFalse(store.read_manifest("stale")["has_pdfs"])
            store.enforce_pdf_cap()                   # ξανά -> no-op, χωρίς σφάλμα
            self.assertFalse(store.read_manifest("stale")["has_pdfs"])

    def test_presigned_url_shape(self):
        """presigned_url -> str που περιέχει το object key και υπογραφή· με
        download_name επιβάλλεται Content-Disposition (URL-encoded). Τα URL
        του moto είναι έγκυρα strings — δεν τα κατεβάζουμε, ελέγχουμε μόνο
        τη μορφή τους (offline)."""
        from urllib.parse import quote
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(Bucket="demo-test-bucket")
            from demolitions.storage import R2Storage
            store = R2Storage("demo-test-bucket", "acc", "k", "s",
                              endpoint_url="https://s3.amazonaws.com")

            url = store.presigned_url("rid1", "rid1.xlsx")
            self.assertIsInstance(url, str)
            self.assertIn("runs/rid1/rid1.xlsx", url)
            self.assertIn("X-Amz-Signature", url)
            # χωρίς download_name -> χωρίς Content-Disposition
            self.assertNotIn("response-content-disposition", url.lower())

            named = store.presigned_url("rid1", "rid1.xlsx", download_name="rid1.xlsx")
            self.assertIsInstance(named, str)
            self.assertIn("X-Amz-Signature", named)
            self.assertIn("response-content-disposition", named.lower())

            # ελληνικό όνομα -> RFC-5987· το όνομα μπαίνει quoted στο
            # Content-Disposition και μετά το URL του query string ξανα-
            # κωδικοποιείται από τον presigner (διπλό encoding)
            greek = store.presigned_url("rid1", "a.pdf",
                                        download_name="Δήμος Χ.zip")
            self.assertIn(quote(quote("Δήμος Χ.zip")), greek)


class TestLocalStorage(unittest.TestCase):
    def test_lifecycle(self):
        from demolitions.storage import LocalStorage
        with tempfile.TemporaryDirectory() as tmp:
            st = LocalStorage(tmp)
            d = st.staging_dir("r1")
            (d / "pdf" / "Δήμος Χ" / "2021").mkdir(parents=True)
            (d / "pdf" / "Δήμος Χ" / "2021" / "a.pdf").write_bytes(b"PDF" * 100)
            (d / "rows.json").write_text("[]", "utf-8")
            (d / "run.json").write_text(
                json.dumps({"run_id": "r1", "has_pdfs": True, "n_rows": 0}), "utf-8")
            (d / "r1.xlsx").write_bytes(b"xlsx")

            self.assertTrue(st.exists("r1"))
            self.assertEqual([m["run_id"] for m in st.list_runs()], ["r1"])
            sizes = st.sizes_by_run()["r1"]
            self.assertEqual(sizes["pdf_bytes"], 300)
            self.assertGreater(sizes["total_bytes"], 300)
            self.assertEqual(st.usage()["pdf_bytes"], 300)
            gen, sz = st.open_member("r1", "r1.xlsx")
            self.assertEqual(b"".join(gen), b"xlsx")
            self.assertEqual(sz, 4)
            self.assertIsNone(st.open_member("r1", "../../etc"))   # traversal
            self.assertEqual(len(list(st.iter_pdfs("r1"))), 1)

            st.delete_pdfs("r1")
            self.assertEqual(list(st.iter_pdfs("r1")), [])
            self.assertTrue(st.exists("r1"))                       # μεταδεδομένα μένουν
            st.delete_run("r1")
            self.assertFalse(st.exists("r1"))

    def test_delete_orphans(self):
        from demolitions.storage import LocalStorage
        with tempfile.TemporaryDirectory() as tmp:
            st = LocalStorage(tmp)
            good = st.staging_dir("good")             # έχει run.json -> μένει
            good.mkdir(parents=True, exist_ok=True)
            (good / "run.json").write_text("{}", "utf-8")
            bad = st.staging_dir("bad")               # χωρίς run.json -> orphan
            (bad / "pdf").mkdir(parents=True)
            (bad / "pdf" / "x.pdf").write_bytes(b"x")
            act = st.staging_dir("active")            # ανεβαίνει ακόμη -> keep
            (act / "pdf").mkdir(parents=True)
            (act / "pdf" / "y.pdf").write_bytes(b"y")

            freed = st.delete_orphans(keep_ids=("active",))
            self.assertEqual(freed, 1)
            self.assertTrue((Path(tmp) / "good").exists())
            self.assertFalse((Path(tmp) / "bad").exists())
            self.assertTrue((Path(tmp) / "active").exists())

    def test_cache_is_noop_local(self):
        # τοπικά το cache_dir είναι ήδη μόνιμο -> pull/push δεν κάνουν τίποτα
        from demolitions.storage import LocalStorage
        with tempfile.TemporaryDirectory() as tmp:
            st = LocalStorage(tmp)
            self.assertFalse(st.pull_cache("geocode.json", tmp))
            self.assertFalse(st.push_cache("geocode.json", tmp))

    def test_presigned_url_is_none_local(self):
        # τοπικά δεν υπάρχει presigning — ο caller σερβίρει το αρχείο
        from demolitions.storage import LocalStorage
        with tempfile.TemporaryDirectory() as tmp:
            st = LocalStorage(tmp)
            self.assertIsNone(st.presigned_url("r1", "r1.xlsx"))
            self.assertIsNone(st.presigned_url("r1", "r1.xlsx",
                                               download_name="r1.xlsx"))


class TestZipManifestR2(unittest.TestCase):
    """api_zip_manifest με R2 backend (moto): client mode όταν υπάρχουν PDF,
    αλλιώς server mode. Το webui.store παρακάμπτεται προσωρινά στον R2 store."""

    RID = "r2-run_2024-01-01_2024-12-31_xyz"

    def setUp(self):
        import importlib.util
        if importlib.util.find_spec("moto") is None:
            self.skipTest("moto δεν είναι εγκατεστημένο")
        # ο rate limiter μοιράζεται το "zip" bucket με τα υπόλοιπα zip tests·
        # καθάρισε ώστε αυτό το test να μην μπλοκαριστεί από προηγούμενα hits
        import webui
        with webui.rate_lock:
            webui.rate_hits.clear()

    def _make_staging(self, store, run_id, n_pdfs, has_pdfs=True):
        d = store.staging_dir(run_id)
        rows = []
        for i in range(n_pdfs):
            rel = f"pdf/Δήμος Χ/2021/ΑΔΑ{i}.pdf"
            p = d / rel
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_bytes(b"%PDF-1.4 test")
            rows.append({"ada": f"ΑΔΑ{i}", "pdf_path": rel})
        (d / "rows.json").write_text(json.dumps(rows, ensure_ascii=False), "utf-8")
        (d / (run_id + ".xlsx")).write_bytes(b"xlsx-bytes")
        (d / "run.json").write_text(json.dumps(
            {"run_id": run_id, "n_rows": n_pdfs, "has_pdfs": has_pdfs,
             "created": "2024-01-01"}), "utf-8")

    def _r2_store(self):
        from demolitions.storage import R2Storage
        return R2Storage("demo-test-bucket", "acc", "k", "s",
                         endpoint_url="https://s3.amazonaws.com")

    def test_client_mode_when_has_pdfs(self):
        import webui
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(
                Bucket="demo-test-bucket")
            store = self._r2_store()
            self._make_staging(store, self.RID, n_pdfs=3)
            store.save_run(self.RID)

            orig = webui.store
            webui.store = store
            try:
                from urllib.parse import quote
                m = webui.app.test_client().get(
                    "/api/runs/" + quote(self.RID) + "/zip-manifest",
                    environ_base={"REMOTE_ADDR": "198.51.100.7"}).get_json()
            finally:
                webui.store = orig

        self.assertEqual(m["mode"], "client")
        self.assertEqual(m["zipname"], self.RID + ".zip")
        self.assertEqual(m["fallback"], "/zip/" + self.RID + ".zip")
        # xlsx + ένα ανά PDF
        self.assertEqual(len(m["files"]), 1 + 3)
        first = m["files"][0]
        self.assertEqual(first["name"], self.RID + ".xlsx")
        self.assertIn("X-Amz-Signature", first["url"])     # presigned
        for f in m["files"]:
            self.assertTrue(f["name"])
            self.assertTrue(f["url"])
            self.assertIn("X-Amz-Signature", f["url"])

    def test_server_mode_when_no_pdfs(self):
        import webui
        from moto import mock_aws
        with mock_aws():
            import boto3
            boto3.client("s3", region_name="us-east-1").create_bucket(
                Bucket="demo-test-bucket")
            store = self._r2_store()
            self._make_staging(store, self.RID, n_pdfs=0, has_pdfs=False)
            store.save_run(self.RID)

            orig = webui.store
            webui.store = store
            try:
                from urllib.parse import quote
                m = webui.app.test_client().get(
                    "/api/runs/" + quote(self.RID) + "/zip-manifest",
                    environ_base={"REMOTE_ADDR": "198.51.100.8"}).get_json()
            finally:
                webui.store = orig

        self.assertEqual(m["mode"], "server")
        self.assertEqual(m["url"], "/zip/" + self.RID + ".zip")


class TestFilteredDownload(unittest.TestCase):
    """«Λήψη φιλτραρισμένων»: POST /api/runs/<id>/xlsx-filtered και το
    φιλτραρισμένο serve_zip μέσω POST (η λίστα ΑΔΑ στο σώμα — `request.form`
    getlist ή JSON). Με φίλτρο → μόνο τα ΑΔΑ του υποσυνόλου· χωρίς φίλτρο (απλό
    GET) η συμπεριφορά είναι byte-for-byte ίδια με πριν."""

    RID = "filt-run_2021-01-01_2022-12-31_xyz"

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp()
        os.environ["DEMOLITIONS_STORAGE"] = "local"
        os.environ["DEMOLITIONS_RUNS_DIR"] = os.path.join(cls.tmp, "runs")
        os.environ["DEMOLITIONS_CACHE_DIR"] = os.path.join(cls.tmp, "cache")
        import importlib
        cls.webui = importlib.reload(importlib.import_module("webui"))
        cls.client = cls.webui.app.test_client()

    @classmethod
    def tearDownClass(cls):
        import shutil
        shutil.rmtree(cls.tmp, ignore_errors=True)
        for k in ("DEMOLITIONS_STORAGE", "DEMOLITIONS_RUNS_DIR",
                  "DEMOLITIONS_CACHE_DIR"):
            os.environ.pop(k, None)

    def setUp(self):
        self.webui.rate_hits.clear()
        self.webui.RATE_LIMIT_WINDOW_SECONDS = 60
        self.webui.RATE_LIMIT_MAX_REQUESTS = 100
        base = {"dim_enotita": "", "poli": "", "odos": "", "ar_apo": "",
                "ar_eos": "", "ot": "", "kaek": "", "perigrafi": "ΚΑΤΕΔΑΦΙΣΗ",
                "orofoi": 1, "lat": None, "lon": None, "precision": "",
                "parse_ok": True, "flags": "", "eidos": "κατεδάφιση"}
        # 3 γραμμές, η μία (Γ) χωρίς PDF· διαφορετικά έτη/δήμοι ώστε το pivot
        # να αλλάζει ορατά στο υποσύνολο
        self.rows = [
            {**base, "ada": "ΑΔΑ-Α", "url": "u/Α", "date": "2021-01-15",
             "year": 2021, "dimos": "Δήμος Α",
             "pdf_path": "pdf/Δήμος Α/2021/ΑΔΑ-Α.pdf"},
            {**base, "ada": "ΑΔΑ:Β", "url": "u/Β", "date": "2021-06-20",
             "year": 2021, "dimos": "Δήμος Β",
             "pdf_path": "pdf/Δήμος Β/2021/ΑΔΑ-Β.pdf"},
            {**base, "ada": "ΑΔΑ-Γ", "url": "u/Γ", "date": "2022-03-10",
             "year": 2022, "dimos": "Δήμος Γ", "pdf_path": ""},
        ]
        d = self.webui.store.staging_dir(self.RID)
        for r in self.rows:
            if r["pdf_path"]:
                p = d / r["pdf_path"]
                p.parent.mkdir(parents=True, exist_ok=True)
                p.write_bytes(b"%PDF-1.4 " + r["ada"].encode("utf-8"))
        write_xlsx(self.rows, d / (self.RID + ".xlsx"))
        (d / "rows.json").write_text(
            json.dumps(self.rows, ensure_ascii=False), "utf-8")
        (d / "run.json").write_text(json.dumps({
            "run_id": self.RID, "area": "Α", "from": "2021-01-01",
            "to": "2022-12-31", "created": "2021-01-15", "n_rows": 3,
            "n_dups": 0, "geocoded": True, "has_pdfs": True},
            ensure_ascii=False), "utf-8")

    def tearDown(self):
        self.webui.store.delete_run(self.RID)

    def _q(self, path):
        from urllib.parse import quote
        return path.replace("<rid>", quote(self.RID))

    # ---- POST /api/runs/<id>/xlsx-filtered ----------------------------------
    def test_xlsx_filtered_subset(self):
        from openpyxl import load_workbook
        # υποσύνολο 2 από 3 ΑΔΑ (συμπεριλ. του «ΑΔΑ:Β» με χαρακτήρα προς encode)
        r = self.client.post(self._q("/api/runs/<rid>/xlsx-filtered"),
                             json={"ada": ["ΑΔΑ-Α", "ΑΔΑ:Β", "ΑΓΝΩΣΤΟ"]})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.headers["Content-Type"],
                         "application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet")
        self.assertIn("attachment", r.headers.get("Content-Disposition", ""))
        wb = load_workbook(io.BytesIO(r.data))
        ws = wb["Κατεδαφίσεις"]
        self.assertEqual(ws.max_row, 3)        # header + 2 γραμμές
        # σειρά διατηρείται από το rows.json (Α πριν Β)
        col = {key: i for i, (_, key, _) in enumerate(COLUMNS, 1)}
        self.assertEqual(ws.cell(2, col["ada"]).value, "ΑΔΑ-Α")
        self.assertEqual(ws.cell(3, col["ada"]).value, "ΑΔΑ:Β")
        # το pivot ξαναϋπολογίστηκε για το υποσύνολο (μόνο 2 άδειες, 2021)
        pivot = wb["Ανά έτος-δήμο"]
        self.assertEqual(pivot.cell(pivot.max_row, pivot.max_column).value, 2)

    def test_xlsx_filtered_unknown_run_404(self):
        r = self.client.post("/api/runs/does-not-exist/xlsx-filtered",
                             json={"ada": ["x"]})
        self.assertEqual(r.status_code, 404)

    def test_xlsx_filtered_bad_input_400(self):
        r = self.client.post(self._q("/api/runs/<rid>/xlsx-filtered"),
                             json={"ada": "not-a-list"})
        self.assertEqual(r.status_code, 400)
        r = self.client.post(self._q("/api/runs/<rid>/xlsx-filtered"),
                             json={"ada": [1, 2]})
        self.assertEqual(r.status_code, 400)

    def test_xlsx_filtered_no_match_400(self):
        r = self.client.post(self._q("/api/runs/<rid>/xlsx-filtered"),
                             json={"ada": ["ΑΓΝΩΣΤΟ"]})
        self.assertEqual(r.status_code, 400)

    # ---- serve_zip με φίλτρο ?ada=… -----------------------------------------
    def test_zip_with_ada_filter(self):
        from openpyxl import load_workbook
        # κράτα μόνο Α (έχει PDF) και Γ (χωρίς PDF) — POST με επαναλαμβανόμενα πεδία
        r = self.client.post(self._q("/zip/<rid>.zip"),
                             data={"ada": ["ΑΔΑ-Α", "ΑΔΑ-Γ"]})
        self.assertEqual(r.status_code, 200)
        zf = zipfile.ZipFile(io.BytesIO(r.data))
        names = zf.namelist()
        pdfs = [n for n in names if n.endswith(".pdf")]
        # μόνο το PDF του Α (το Γ δεν έχει pdf_path, το Β δεν είναι στο φίλτρο)
        self.assertEqual(pdfs, ["pdf/Δήμος Α/2021/ΑΔΑ-Α.pdf"])
        # το xlsx ξαναφτιάχτηκε για το υποσύνολο (Α + Γ = 2 γραμμές)
        xlsx = [n for n in names if n.endswith(".xlsx")]
        self.assertEqual(len(xlsx), 1)
        wb = load_workbook(io.BytesIO(zf.read(xlsx[0])))
        self.assertEqual(wb["Κατεδαφίσεις"].max_row, 3)   # header + Α + Γ

    def test_zip_without_filter_unchanged(self):
        # χωρίς ?ada= → πλήρες zip (xlsx + όλα τα PDF), αμετάβλητο
        r = self.client.get(self._q("/zip/<rid>.zip"))
        self.assertEqual(r.status_code, 200)
        zf = zipfile.ZipFile(io.BytesIO(r.data))
        names = zf.namelist()
        pdfs = sorted(n for n in names if n.endswith(".pdf"))
        self.assertEqual(pdfs, ["pdf/Δήμος Α/2021/ΑΔΑ-Α.pdf",
                                "pdf/Δήμος Β/2021/ΑΔΑ-Β.pdf"])
        xlsx = [n for n in names if n.endswith(".xlsx")]
        self.assertEqual(xlsx, [self.RID + ".xlsx"])
        # το πλήρες xlsx έχει 3 γραμμές
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(zf.read(xlsx[0])))
        self.assertEqual(wb["Κατεδαφίσεις"].max_row, 4)   # header + 3


class TestWebUI(unittest.TestCase):
    """Ολοκληρωμένος έλεγχος των endpoints με τοπικό store σε temp φάκελο."""

    RID = "test-run_2021-01-01_2021-12-31_xyz"

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp()
        os.environ["DEMOLITIONS_STORAGE"] = "local"
        os.environ["DEMOLITIONS_RUNS_DIR"] = os.path.join(cls.tmp, "runs")
        os.environ["DEMOLITIONS_CACHE_DIR"] = os.path.join(cls.tmp, "cache")
        import importlib
        cls.webui = importlib.reload(importlib.import_module("webui"))
        cls.client = cls.webui.app.test_client()

    @classmethod
    def tearDownClass(cls):
        import shutil
        shutil.rmtree(cls.tmp, ignore_errors=True)
        for k in ("DEMOLITIONS_STORAGE", "DEMOLITIONS_RUNS_DIR", "DEMOLITIONS_CACHE_DIR"):
            os.environ.pop(k, None)

    def setUp(self):
        # φρέσκο ολοκληρωμένο run πριν από κάθε test (κάποια το διαγράφουν)
        from demolitions.output import write_xlsx
        self.webui.rate_hits.clear()
        self.webui.RATE_LIMIT_WINDOW_SECONDS = 60
        self.webui.RATE_LIMIT_MAX_REQUESTS = 12
        d = self.webui.store.staging_dir(self.RID)
        (d / "pdf" / "Δήμος Χ" / "2021").mkdir(parents=True, exist_ok=True)
        (d / "pdf" / "Δήμος Χ" / "2021" / "ΑΔΑ1.pdf").write_bytes(b"%PDF-1.4 x")
        rows = [{
            "ada": "ΑΔΑ1", "url": "https://diavgeia.gov.gr/decision/view/ΑΔΑ1",
            "date": "2021-01-15", "year": 2021, "dimos": "Δήμος Χ",
            "eidos": "κατεδάφιση", "ektasi": "ολική", "dim_enotita": "",
            "poli": "Χ", "odos": "Οδός", "ar_apo": "1", "ar_eos": "", "ot": "",
            "kaek": "", "perigrafi": "ΚΑΤΕΔΑΦΙΣΗ", "orofoi": 1,
            "lat": 41.1, "lon": 24.1, "precision": "κτίσμα (PDF)",
            "parse_ok": True, "pdf_path": "pdf/Δήμος Χ/2021/ΑΔΑ1.pdf",
            "flags": "", "poly": [[41.1, 24.1], [41.1, 24.11], [41.11, 24.11]],
        }]
        write_xlsx(rows, d / (self.RID + ".xlsx"))
        (d / "rows.json").write_text(json.dumps(rows, ensure_ascii=False), "utf-8")
        (d / "run.json").write_text(json.dumps({
            "run_id": self.RID, "area": "Δήμος Χ", "from": "2021-01-01",
            "to": "2021-12-31", "created": "2021-01-15", "n_rows": 1,
            "n_dups": 0, "geocoded": True, "has_pdfs": True}, ensure_ascii=False), "utf-8")

    def _url(self, path):
        from urllib.parse import quote
        return path.replace("<rid>", quote(self.RID))

    def test_index_and_areas(self):
        r = self.client.get("/")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"tab-about", r.data)
        self.assertIn(b'value="2018-10-01"', r.data)   # default «Από»
        self.assertGreater(len(self.client.get("/api/areas").get_json()), 300)

    def test_download_links_derived_from_run_id(self):
        # zip/xlsx σύνδεσμοι χτίζονται από το run_id (δουλεύουν και στο
        # ιστορικό)· όχι από manifest.zip_url που λείπει εκεί -> «/undefined»
        html = self.client.get("/").get_data(as_text=True)
        self.assertIn("/zip/${rid}.zip", html)
        self.assertIn("/runs/${rid}/${rid}.xlsx", html)
        self.assertNotIn("${manifest.zip_url}", html)
        self.assertNotIn("${manifest.xlsx_url}", html)

    def test_healthz(self):
        r = self.client.get("/healthz")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data, b"ok")

    def test_about(self):
        a = self.client.get("/api/about").get_json()
        self.assertEqual(a["backend"], "local")
        self.assertGreaterEqual(a["n_runs"], 1)
        self.assertEqual(a["data_since"], "2018-10-01")

    def test_runs_has_sizes(self):
        m = next(x for x in self.client.get("/api/runs").get_json()
                 if x["run_id"] == self.RID)
        self.assertGreater(m["total_bytes"], 0)
        self.assertGreater(m["pdf_bytes"], 0)

    def test_serve_xlsx_and_rows(self):
        r = self.client.get(self._url("/runs/<rid>/<rid>.xlsx".replace(
            "<rid>.xlsx", __import__("urllib.parse", fromlist=["quote"]).quote(self.RID) + ".xlsx")))
        self.assertEqual(r.status_code, 200)
        self.assertIn("attachment", r.headers.get("Content-Disposition", ""))
        rows = self.client.get(self._url("/api/runs/<rid>/rows")).get_json()
        self.assertEqual(len(rows), 1)

    def test_zip_contains_xlsx_and_pdf(self):
        r = self.client.get(self._url("/zip/<rid>.zip"))
        self.assertEqual(r.status_code, 200)
        names = zipfile.ZipFile(io.BytesIO(r.data)).namelist()
        self.assertTrue(any(n.endswith(".xlsx") for n in names))
        self.assertTrue(any(n.endswith(".pdf") for n in names))

    def test_zip_opens_members_lazily(self):
        # τα μέλη (xlsx/PDF) ΔΕΝ πρέπει να ανοίγονται κατά την κατασκευή της
        # απόκρισης — αλλιώς N round-trips πριν το πρώτο byte (-> 502 hosted)
        w = self.webui
        calls = []
        orig = w.store.open_member
        w.store.open_member = lambda rid, rel: (calls.append(rel) or orig(rid, rel))
        try:
            with w.app.test_request_context():
                resp = w.serve_zip(self.RID)
                self.assertEqual(calls, ["rows.json"])   # μόνο αυτό κατά την κατασκευή
                body = b"".join(resp.response)            # τώρα streamάρει
            self.assertIn(self.RID + ".xlsx", calls)      # ανοίχτηκε κατά το stream
            self.assertTrue(any(c.endswith(".pdf") for c in calls))
            self.assertTrue(zipfile.ZipFile(io.BytesIO(body)).namelist())
        finally:
            w.store.open_member = orig

    def test_zip_falls_back_to_diavgeia_when_cached_pdf_missing(self):
        """has_pdfs=True αλλά λείπει αντικείμενο στο R2 -> το zip το παίρνει
        από τη Διαύγεια αντί για 0-byte εγγραφή (μικρότερο zip)."""
        from unittest import mock
        w = self.webui
        orig = w.store.open_member

        def fake_open(rid, rel):
            if rel.endswith(".pdf"):
                return None                  # cached PDF λείπει
            return orig(rid, rel)

        with mock.patch.object(w.store, "open_member", side_effect=fake_open), \
             mock.patch.object(w, "_diavgeia_pdf",
                               side_effect=lambda ada: iter([b"%PDF-FALLBACK"])):
            with w.app.test_request_context():
                resp = w.serve_zip(self.RID)
                body = b"".join(resp.response)
        zf = zipfile.ZipFile(io.BytesIO(body))
        pdfs = [n for n in zf.namelist() if n.endswith(".pdf")]
        self.assertTrue(pdfs)
        self.assertEqual(zf.read(pdfs[0]), b"%PDF-FALLBACK")   # όχι 0 bytes

    def test_zip_rate_limit(self):
        self.webui.RATE_LIMIT_WINDOW_SECONDS = 3600
        self.webui.RATE_LIMIT_MAX_REQUESTS = 1
        r1 = self.client.get(self._url("/zip/<rid>.zip"),
                             environ_base={"REMOTE_ADDR": "203.0.113.9"})
        self.assertEqual(r1.status_code, 200)
        r2 = self.client.get(self._url("/zip/<rid>.zip"),
                             environ_base={"REMOTE_ADDR": "203.0.113.9"})
        self.assertEqual(r2.status_code, 429)
        self.assertIn("Retry-After", r2.headers)

    def test_run_validation(self):
        r = self.client.post("/api/run", json={
            "area": "Δήμος Ασγκαμπάτ", "from": "2021-01-01", "to": "2021-12-31"})
        self.assertEqual(r.status_code, 400)
        r = self.client.post("/api/run", json={
            "area": "Δήμος Δράμας", "from": "2022-01-01", "to": "2021-01-01"})
        self.assertEqual(r.status_code, 400)

    def test_traversal_blocked(self):
        r = self.client.get(self._url("/runs/<rid>/..%2f..%2fwebui.py"))
        self.assertIn(r.status_code, (400, 404))

    def test_clear_pdfs_then_delete(self):
        r = self.client.delete(self._url("/api/runs/<rid>/pdfs"))
        self.assertEqual(r.status_code, 200)
        m = next(x for x in self.client.get("/api/runs").get_json()
                 if x["run_id"] == self.RID)
        self.assertFalse(m["has_pdfs"])
        self.assertEqual(m["pdf_bytes"], 0)
        self.assertTrue(self.webui.store.exists(self.RID))   # μεταδεδομένα μένουν
        r = self.client.delete(self._url("/api/runs/<rid>"))
        self.assertEqual(r.status_code, 200)
        self.assertFalse(self.webui.store.exists(self.RID))

    def test_delete_all(self):
        self.assertTrue(self.webui.store.exists(self.RID))
        r = self.client.delete("/api/runs")
        self.assertEqual(r.status_code, 200)
        self.assertGreaterEqual(r.get_json()["deleted"], 1)
        self.assertEqual(self.client.get("/api/runs").get_json(), [])

    def test_delete_all_pdfs(self):
        r = self.client.delete("/api/pdfs")
        self.assertEqual(r.status_code, 200)
        self.assertGreaterEqual(r.get_json()["cleared"], 1)
        m = next(x for x in self.client.get("/api/runs").get_json()
                 if x["run_id"] == self.RID)
        self.assertFalse(m["has_pdfs"])           # PDF καθαρίστηκαν
        self.assertEqual(m["pdf_bytes"], 0)
        self.assertTrue(self.webui.store.exists(self.RID))   # το run μένει

    def test_keepalive_noop_without_url(self):
        # τοπικά (χωρίς RENDER_EXTERNAL_URL) το keep-alive δεν κάνει αίτημα
        from unittest import mock
        w = self.webui
        with mock.patch.object(w, "KEEPALIVE_URL", None), \
             mock.patch.object(w.urllib.request, "urlopen") as urlopen:
            j = w.Job()
            j.state = "running"
            w._keepalive_loop(j)        # επιστρέφει αμέσως
            urlopen.assert_not_called()

    def test_keepalive_pings_then_stops_when_job_done(self):
        from unittest import mock
        w = self.webui
        j = w.Job()
        j.state = "running"

        # στο πρώτο ping το job «τελειώνει» -> ο βρόχος σταματά αμέσως μετά
        def stop(*a, **k):
            j.state = "done"
            return mock.MagicMock()

        with mock.patch.object(w, "KEEPALIVE_URL", "https://x.onrender.com"), \
             mock.patch.object(w.urllib.request, "urlopen",
                               side_effect=stop) as urlopen, \
             mock.patch.object(w.time, "sleep"):
            w._keepalive_loop(j)
        urlopen.assert_called_once()
        self.assertIn("/healthz", urlopen.call_args[0][0])

    def test_delete_all_pdfs_sweeps_orphans(self):
        # φάκελος run χωρίς run.json (ημιτελές) -> τον καθαρίζει το /api/pdfs
        orphan = self.webui.store.staging_dir("orphan-run")
        (orphan / "pdf").mkdir(parents=True, exist_ok=True)
        (orphan / "pdf" / "z.pdf").write_bytes(b"%PDF x")
        self.assertTrue(orphan.exists())
        r = self.client.delete("/api/pdfs")
        self.assertEqual(r.status_code, 200)
        self.assertGreaterEqual(r.get_json().get("orphans", 0), 1)
        self.assertFalse(orphan.exists())
        self.assertTrue(self.webui.store.exists(self.RID))   # ολοκληρωμένο μένει

    def test_rate_limit_blocks_after_max(self):
        w = self.webui
        with w.rate_lock:
            w.rate_hits.clear()
        with w.app.test_request_context(environ_base={"REMOTE_ADDR": "7.7.7.7"}):
            results = [w._check_rate_limit("zip")
                       for _ in range(w.RATE_LIMIT_MAX_REQUESTS + 2)]
        self.assertIsNone(results[0])
        self.assertIsNotNone(results[-1])          # μπλοκάρεται μετά το όριο

    def test_rate_limit_prunes_stale_keys(self):
        w = self.webui
        with w.rate_lock:
            w.rate_hits.clear()
            w.rate_hits[("1.1.1.1", "run")] = [0.0]   # έληξε (timestamp 0)
            w.rate_hits[("2.2.2.2", "zip")] = []        # κενή
        with w.app.test_request_context(environ_base={"REMOTE_ADDR": "9.9.9.9"}):
            w._check_rate_limit("run")
        self.assertNotIn(("1.1.1.1", "run"), w.rate_hits)
        self.assertNotIn(("2.2.2.2", "zip"), w.rate_hits)
        self.assertIn(("9.9.9.9", "run"), w.rate_hits)

    def test_zip_manifest_server_mode_local(self):
        # τοπικός store -> πάντα server-side zip (δεν υπάρχει presigning)
        m = self.client.get(self._url("/api/runs/<rid>/zip-manifest")).get_json()
        self.assertEqual(m["mode"], "server")
        self.assertEqual(m["url"], "/zip/{}.zip".format(self.RID))

    def test_zip_manifest_unknown_run_404(self):
        r = self.client.get("/api/runs/does-not-exist/zip-manifest")
        self.assertEqual(r.status_code, 404)

    def test_index_frontend_wiring(self):
        # η σελίδα φέρνει όλα τα σημεία σύνδεσης του client-side zip download
        html = self.client.get("/").get_data(as_text=True)
        self.assertIn("data-r2", html)                       # <body data-r2=…>
        self.assertIn('document.body.dataset.r2 === "true"', html)
        self.assertIn("/static/client-zip.js", html)
        self.assertIn("data-zip-run", html)                  # κουμπί client-zip
        self.assertIn("data-zip-filtered", html)             # κουμπί φιλτραρισμένων
        self.assertIn("/xlsx-filtered", html)                # POST endpoint
        self.assertIn("__demolitionsFilter", html)           # γέφυρα classic→module
        self.assertIn('type="module"', html)                 # ESM import


if __name__ == "__main__":
    unittest.main(verbosity=2)
