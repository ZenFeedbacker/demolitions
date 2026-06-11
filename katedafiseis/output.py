"""Εξαγωγή σε xlsx: ένα φύλλο με τις κατεδαφίσεις, ένα pivot ανά έτος/δήμο."""

from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("ΑΔΑ", "ada", 16),
    ("Ημ/νία έκδοσης", "date", 14),
    ("Έτος", "year", 7),
    ("Δήμος", "dimos", 26),
    ("Δημοτική Ενότητα/Περιοχή", "dim_enotita", 24),
    ("Πόλη/Οικισμός", "poli", 20),
    ("Οδός", "odos", 24),
    ("Αρ. από", "ar_apo", 9),
    ("Αρ. έως", "ar_eos", 9),
    ("ΟΤ", "ot", 8),
    ("ΚΑΕΚ", "kaek", 16),
    ("Περιγραφή έργου", "perigrafi", 60),
    ("Όροφοι", "orofoi", 8),
    ("Lat", "lat", 11),
    ("Lon", "lon", 11),
    ("Ακρίβεια συντ/νων", "precision", 14),
    ("Ανάλυση PDF", "parse_ok", 12),
    ("PDF", "pdf_path", 8),
    ("Έλεγχος", "flags", 24),
]


def write_xlsx(rows, out_path):
    wb = Workbook()

    ws = wb.active
    ws.title = "Κατεδαφίσεις"
    bold = Font(bold=True)
    for col, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
        ws.column_dimensions[get_column_letter(col)].width = width
    for r, row in enumerate(rows, 2):
        for col, (_, key, _) in enumerate(COLUMNS, 1):
            value = row.get(key)
            if key == "parse_ok":
                value = "ΝΑΙ" if value else "ΟΧΙ"
            if key == "pdf_path":
                cell = ws.cell(row=r, column=col,
                               value="PDF" if value else "")
                if value:
                    cell.hyperlink = value  # σχετική διαδρομή δίπλα στο xlsx
                    cell.font = Font(color="0563C1", underline="single")
                continue
            cell = ws.cell(row=r, column=col, value=value)
            if key == "ada":
                cell.hyperlink = row["url"]
                cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 2)}"

    # pivot: γραμμές = έτη, στήλες = δήμοι
    pivot = Counter((row["year"], row["dimos"]) for row in rows)
    years = sorted({y for y, _ in pivot})
    dimoi = sorted({d for _, d in pivot})
    ws2 = wb.create_sheet("Ανά έτος-δήμο")
    ws2.cell(row=1, column=1, value="Έτος").font = bold
    for c, d in enumerate(dimoi, 2):
        ws2.cell(row=1, column=c, value=d).font = bold
        ws2.column_dimensions[get_column_letter(c)].width = max(14, len(d) + 2)
    ws2.cell(row=1, column=len(dimoi) + 2, value="Σύνολο").font = bold
    for r, y in enumerate(years, 2):
        ws2.cell(row=r, column=1, value=y).font = bold
        for c, d in enumerate(dimoi, 2):
            ws2.cell(row=r, column=c, value=pivot.get((y, d), 0))
        ws2.cell(row=r, column=len(dimoi) + 2,
                 value=sum(pivot.get((y, d), 0) for d in dimoi))
    total_row = len(years) + 2
    ws2.cell(row=total_row, column=1, value="Σύνολο").font = bold
    for c, d in enumerate(dimoi, 2):
        ws2.cell(row=total_row, column=c,
                 value=sum(pivot.get((y, d), 0) for y in years))
    ws2.cell(row=total_row, column=len(dimoi) + 2, value=len(rows)).font = bold

    wb.save(out_path)
